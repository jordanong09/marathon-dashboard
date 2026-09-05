import base64
import re
from io import BytesIO, StringIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =========================================================
# PAGE CONFIGURATION
# =========================================================

st.set_page_config(
    page_title="BYD Singapore International Marathon — Registration Intelligence",
    page_icon="🏃",
    layout="wide",
)


# =========================================================
# EVENT BANNER
# =========================================================
# The official event logo file is expected next to app.py. If it
# is missing (for example, only app.py was copied into an old
# bundle), the dashboard falls back to a plain text title rather
# than breaking.

LOGO_FILENAME = "event_logo.png"

# The lion red from the event mark, used for the banner rule.
BRAND_RED = "#E8402D"


def render_event_banner():
    """
    Render a professional header: event logo on the left, report
    title and subtitle on the right, separated from the content by
    a brand-red rule. Returns True when the logo was found.
    """
    logo_path = Path(__file__).parent / LOGO_FILENAME

    if not logo_path.exists():
        return False

    encoded_logo = base64.b64encode(
        logo_path.read_bytes()
    ).decode()

    st.markdown(
        f"""
        <div style="
            display: flex;
            align-items: center;
            gap: 1.25rem;
            padding: 0.4rem 0 0.9rem 0;
            border-bottom: 3px solid {BRAND_RED};
            margin-bottom: 1.1rem;
            flex-wrap: wrap;
        ">
            <img src="data:image/png;base64,{encoded_logo}"
                 alt="BYD Singapore International Marathon"
                 style="height: 74px;" />
            <div>
                <div style="
                    font-size: 1.55rem;
                    font-weight: 700;
                    line-height: 1.2;
                    color: #31333F;
                ">
                    Registration Intelligence Dashboard
                </div>
                <div style="
                    font-size: 0.9rem;
                    color: #6B7280;
                    margin-top: 0.15rem;
                ">
                    Race weekend 4&ndash;6 Dec 2026 &middot;
                    Registration closes 30 Sep 2026 &middot;
                    Internal management report
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    return True


if not render_event_banner():
    st.title(
        "BYD Singapore International Marathon — "
        "Registration Intelligence Dashboard"
    )

st.caption(
    "Upload the latest registration CSV file to analyse registration pace, "
    "race categories, participant demographics and country representation."
)


# =========================================================
# CONSTANTS
# =========================================================

CATEGORY_ORDER = [
    "BYD Marathon",
    "adidas Half Marathon",
    "Standard Chartered 10km",
    "5km",
    "Kids Dash Competitive 1.6KM",
    "Kids Dash Non-Competitive 1.6KM",
    "Kids Dash Non-Competitive 600m",
    "BYD Marathon Crew Challenge",
]

AGE_GROUP_ORDER = [
    "0-18",
    "19-40",
    "41-55",
    "Above 55",
    "Unknown / Invalid",
]

WEEKDAY_ORDER = [
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
]

SINGAPORE_ALIASES = {
    "singapore",
    "sg",
    "sgp",
    "singaporean",
    "republic of singapore",
}

COLUMN_ALIASES = {
    "registration_date": [
        "registration date",
        "registration datetime",
        "registration date time",
        "registration time",
        "created date",
        "created at",
        "date registered",
        "order date",
    ],
    "age": [
        "current age",
        "age",
        "participant age",
    ],
    "gender": [
        "gender",
        "sex",
        "participant gender",
    ],
    "country": [
        "country",
        "country of residence",
        "residence country",
        "participant country",
    ],
    "nationality": [
        "nationality",
        "citizenship",
    ],
    "category": [
        "category name",
        "category",
        "race category",
        "event category",
    ],
    "group_corporate": [
        "group/corporate name",
        "group corporate name",
        "group / corporate name",
        "corporate name",
        "group name",
    ],
    "addon_bag": [
        "gear deposit bag",
        "deposit bag",
    ],
    "addon_itab": [
        "itab",
    ],
    "addon_runflex": [
        "runflex",
    ],
}

# Friendly labels for the add-on columns, used throughout the
# Corporate, Comps & Add-Ons tab and the Excel export.
ADDON_LABELS = {
    "addon_bag": "Gear Deposit Bag",
    "addon_itab": "iTab",
    "addon_runflex": "RunFlex",
}


# =========================================================
# EVENT CONFIGURATION  (edit these values each season)
# =========================================================

# Race weekend is 4-6 Dec 2026. The split below is an assumption:
# Kids Dash on Friday, 5km/10km on Saturday, Marathon distances on
# Sunday. Adjust the dates if the official programme differs.
RACE_DATES = {
    "BYD Marathon": pd.Timestamp("2026-12-06"),
    "adidas Half Marathon": pd.Timestamp("2026-12-06"),
    "Standard Chartered 10km": pd.Timestamp("2026-12-05"),
    "5km": pd.Timestamp("2026-12-05"),
    "Kids Dash Competitive 1.6KM": pd.Timestamp("2026-12-04"),
    "Kids Dash Non-Competitive 1.6KM": pd.Timestamp("2026-12-04"),
    "Kids Dash Non-Competitive 600m": pd.Timestamp("2026-12-04"),
    "BYD Marathon Crew Challenge": pd.Timestamp("2026-12-06"),
}

# Registration closes before race day. All pacing projections run to
# this date, not to race day.
REGISTRATION_CLOSE_DATE = pd.Timestamp("2026-09-30")

# Registration opened on this date. The launch-day surge is a
# structural one-off, so it is excluded from medians, "best day"
# call-outs and pattern statistics — otherwise every chart would
# just report "launch day was the highest", which tells an
# executive nothing.
REGISTRATION_OPEN_DATE = pd.Timestamp("2026-04-27")

LAUNCH_COLOR = "#7B2D8B"       # Launch-day bar (known outlier)

# Targets are defined per "target group". The two Kids Dash 1.6KM
# categories share one combined target of 1,500 (per management
# guidance). The Crew Challenge has no registration target.
TARGET_GROUPS = {
    "BYD Marathon": [
        "BYD Marathon",
    ],
    "adidas Half Marathon": [
        "adidas Half Marathon",
    ],
    "Standard Chartered 10km": [
        "Standard Chartered 10km",
    ],
    "5km": [
        "5km",
    ],
    "Kids Dash 1.6KM (combined)": [
        "Kids Dash Competitive 1.6KM",
        "Kids Dash Non-Competitive 1.6KM",
    ],
    "Kids Dash 600m": [
        "Kids Dash Non-Competitive 600m",
    ],
}

CATEGORY_TARGETS = {
    "BYD Marathon": 13000,
    "adidas Half Marathon": 19000,
    "Standard Chartered 10km": 9000,
    "5km": 7000,
    "Kids Dash 1.6KM (combined)": 1500,
    "Kids Dash 600m": 2500,
}

# Campaign milestones are drawn as dotted vertical markers on every
# time-series chart so registration spikes explain themselves.
# Empty is fine (nothing is drawn) — as a first-year event there is
# no history yet, but add entries the moment a price-tier deadline
# or marketing push is scheduled, as "YYYY-MM-DD": "Label".
MILESTONES = {
    # "2026-08-01": "Early bird ends",
}

# The latest date in an export is usually a partial day, so it is
# excluded from run-rate and projection calculations by default.
EXCLUDE_LATEST_DAY_FROM_PACE = True

# Week-on-week percentage changes are hidden when the previous-week
# base is below this count, because tiny bases produce misleading
# swings (e.g. 2 -> 4 registrations shows as +100%).
SMALL_BASE_THRESHOLD = 30


# =========================================================
# CORPORATE AND COMPLIMENTARY CHANNEL PARSING
# =========================================================
# The Group/Corporate Name column mixes two identities in one
# field, distinguished by prefix:
#   COMPLIMENTARY_<programme>            e.g. COMPLIMENTARY_KOL
#   GROUP_REGISTRATION_<company details> e.g. GROUP_REGISTRATION_
#       SCB Batch 1 - Tan, Hui Hoon
# Rows with neither prefix (or an empty cell) are ordinary public
# registrations.

COMPLIMENTARY_PREFIX = "COMPLIMENTARY"
GROUP_REGISTRATION_PREFIX = "GROUP_REGISTRATION"

# Some companies register under more than one label. Fold known
# variants into a single reporting name so their category mix is
# counted together. Keys are compared after underscores become
# spaces.
COMPANY_ALIASES = {
    "SCB": "Standard Chartered",
    "Standard Chartered Staff": "Standard Chartered",
}

REGISTRATION_TYPE_ORDER = [
    "Public",
    "Group Registration",
    "Complimentary",
]


# =========================================================
# COLOUR SYSTEM
# =========================================================
# Principle: grey is the default, colour is the message.
# Categories keep one fixed colour across every chart and tab.

CATEGORY_COLORS = {
    "BYD Marathon": "#1F4E79",
    "adidas Half Marathon": "#2E86AB",
    "Standard Chartered 10km": "#5AA9C9",
    "5km": "#63A375",
    "Kids Dash Competitive 1.6KM": "#E8871E",
    "Kids Dash Non-Competitive 1.6KM": "#F4B266",
    "Kids Dash Non-Competitive 600m": "#C9A227",
    "BYD Marathon Crew Challenge": "#7B2D8B",
}

ACCENT_COLOR = "#1F4E79"       # Highlight / headline series
NEUTRAL_COLOR = "#B8C4CE"      # De-emphasised bars
UNKNOWN_COLOR = "#9AA5AE"      # Always used for Unknown values
POSITIVE_COLOR = "#2E7D32"     # Increases
NEGATIVE_COLOR = "#C0392B"     # Decreases
PROJECTION_COLOR = "#8A8F98"   # Dashed projection line
TARGET_COLOR = "#C0392B"       # Target reference line

# Age groups are ordinal, so they use a sequential single-hue ramp
# (light = young, dark = old) rather than a qualitative palette.
AGE_GROUP_COLORS = {
    "0-18": "#DCE9F5",
    "19-40": "#8AB4DB",
    "41-55": "#3D7ABC",
    "Above 55": "#16406E",
    "Unknown / Invalid": UNKNOWN_COLOR,
}

GENDER_COLORS = {
    "Male": "#2E86AB",
    "Female": "#E8871E",
    "Other": "#7B2D8B",
    "Unknown": UNKNOWN_COLOR,
}

MARKET_COLORS = {
    "Singapore": "#C8102E",
    "International": "#2E86AB",
    "Unknown": UNKNOWN_COLOR,
}

STATUS_COLORS = {
    "Target met": "#1B5E20",
    "On track": POSITIVE_COLOR,
    "At risk": "#E8871E",
    "Off track": NEGATIVE_COLOR,
}

REGISTRATION_TYPE_COLORS = {
    "Public": NEUTRAL_COLOR,
    "Group Registration": ACCENT_COLOR,
    "Complimentary": "#7B2D8B",
    "Other": UNKNOWN_COLOR,
}


# =========================================================
# TEXT CLEANING AND NORMALISATION
# =========================================================

def normalise_text(value):
    """
    Convert a value into lower-case, consistently spaced text.
    """
    if pd.isna(value):
        return ""

    text = str(value).strip().lower()

    text = text.replace("–", "-")
    text = text.replace("—", "-")
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text)

    return text


def clean_display_text(value):
    """
    Clean text while retaining readable capitalisation.
    """
    if pd.isna(value):
        return "Unknown"

    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)

    if not text:
        return "Unknown"

    return text


# =========================================================
# CATEGORY GROUPING
# =========================================================

def assign_grouped_category(category_name):
    """
    Consolidate detailed race categories into the eight reporting groups.
    """
    category = normalise_text(category_name)

    if not category:
        return "Unmapped"

    # Specific categories must be checked before broad marathon rules.
    if "crew challenge" in category:
        return "BYD Marathon Crew Challenge"

    if (
        "kids dash" in category
        and (
            "non-competitive" in category
            or "non competitive" in category
        )
        and re.search(r"\b600\s*m\b", category)
    ):
        return "Kids Dash Non-Competitive 600m"

    if (
        "kids dash" in category
        and (
            "non-competitive" in category
            or "non competitive" in category
        )
        and re.search(r"\b1\.6\s*km\b", category)
    ):
        return "Kids Dash Non-Competitive 1.6KM"

    if (
        "kids dash" in category
        and "competitive" in category
        and "non-competitive" not in category
        and "non competitive" not in category
        and re.search(r"\b1\.6\s*km\b", category)
    ):
        return "Kids Dash Competitive 1.6KM"

    if "half marathon" in category:
        return "adidas Half Marathon"

    if re.search(r"\b10\s*km\b", category):
        return "Standard Chartered 10km"

    if re.search(r"\b5\s*km\b", category):
        return "5km"

    if "marathon" in category:
        return "BYD Marathon"

    return "Unmapped"


# =========================================================
# GENDER CLEANING
# =========================================================

def normalise_gender(value):
    """
    Consolidate common gender variations.
    """
    gender = normalise_text(value)

    if gender in {"male", "m", "man"}:
        return "Male"

    if gender in {"female", "f", "woman"}:
        return "Female"

    if gender in {
        "non-binary",
        "non binary",
        "nonbinary",
        "others",
        "other",
    }:
        return "Other"

    if not gender:
        return "Unknown"

    return clean_display_text(value)


# =========================================================
# COUNTRY CLEANING
# =========================================================

def normalise_country(value):
    """
    Clean country text and consolidate common Singapore variations.
    """
    country = normalise_text(value)

    if not country:
        return "Unknown"

    if country in SINGAPORE_ALIASES:
        return "Singapore"

    # Title case provides consistent display for most country names.
    return country.title()


def classify_market(country):
    """
    Classify participants as Singapore, International or Unknown.
    """
    if country == "Singapore":
        return "Singapore"

    if country == "Unknown":
        return "Unknown"

    return "International"


# =========================================================
# AGE PROCESSING
# =========================================================

def assign_age_group(age):
    """
    Assign numeric age into a management reporting age band.
    """
    if pd.isna(age):
        return "Unknown / Invalid"

    if age < 0 or age > 110:
        return "Unknown / Invalid"

    if age <= 18:
        return "0-18"

    if age <= 40:
        return "19-40"

    if age <= 55:
        return "41-55"

    return "Above 55"


# =========================================================
# FILE READING
# =========================================================

def read_csv_file(uploaded_file):
    """
    Read an uploaded CSV using common encodings and delimiter detection.
    """
    encodings = [
        "utf-8-sig",
        "utf-8",
        "cp1252",
        "latin-1",
    ]

    last_error = None

    for encoding in encodings:
        try:
            uploaded_file.seek(0)

            dataframe = pd.read_csv(
                uploaded_file,
                encoding=encoding,
                sep=None,
                engine="python"
            )

            dataframe.columns = [
                str(column).strip()
                for column in dataframe.columns
            ]

            return dataframe

        except UnicodeDecodeError as error:
            last_error = error

        except pd.errors.ParserError as error:
            last_error = error

    raise ValueError(
        "The CSV file could not be read. Check its encoding and delimiter."
    ) from last_error


# =========================================================
# COLUMN DETECTION
# =========================================================

def detect_column(columns, aliases):
    """
    Find the best matching source column based on known aliases.
    """
    normalised_columns = {
        normalise_text(column): column
        for column in columns
    }

    # First try an exact match.
    for alias in aliases:
        if alias in normalised_columns:
            return normalised_columns[alias]

    # Then try partial matching.
    for column in columns:
        normalised_column = normalise_text(column)

        for alias in aliases:
            if alias in normalised_column:
                return column

    return None


def optional_column_selector(
    label,
    columns,
    detected_column=None,
    help_text=None,
):
    """
    Create a select box allowing an optional 'Not available' selection.
    """
    options = ["Not available"] + list(columns)

    if detected_column in columns:
        default_index = options.index(detected_column)
    else:
        default_index = 0

    selected = st.selectbox(
        label,
        options=options,
        index=default_index,
        help=help_text,
    )

    if selected == "Not available":
        return None

    return selected


# =========================================================
# DATA PREPARATION
# =========================================================

def prepare_registration_data(
    dataframe,
    registration_date_column,
    category_column,
    age_column=None,
    gender_column=None,
    country_column=None,
    nationality_column=None,
    group_corporate_column=None,
    addon_columns=None,
    day_first=True,
):
    """
    Clean and prepare the uploaded data for dashboard analysis.
    """
    data = dataframe.copy()

    data["Registration Date Time"] = pd.to_datetime(
        data[registration_date_column],
        errors="coerce",
        dayfirst=day_first,
    )

    data["Registration Date Only"] = (
        data["Registration Date Time"]
        .dt.normalize()
    )

    data["Registration Hour"] = (
        data["Registration Date Time"]
        .dt.hour
    )

    data["Registration Weekday"] = (
        data["Registration Date Time"]
        .dt.day_name()
    )

    data["Original Category"] = (
        data[category_column]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    data["Grouped Category"] = (
        data["Original Category"]
        .apply(assign_grouped_category)
    )

    if age_column:
        data["Current Age Clean"] = pd.to_numeric(
            data[age_column],
            errors="coerce",
        )
    else:
        data["Current Age Clean"] = pd.NA

    data["Age Group"] = (
        data["Current Age Clean"]
        .apply(assign_age_group)
    )

    if gender_column:
        data["Gender Clean"] = (
            data[gender_column]
            .apply(normalise_gender)
        )
    else:
        data["Gender Clean"] = "Unknown"

    if country_column:
        data["Country Clean"] = (
            data[country_column]
            .apply(normalise_country)
        )
    else:
        data["Country Clean"] = "Unknown"

    data["Market"] = (
        data["Country Clean"]
        .apply(classify_market)
    )

    if nationality_column:
        data["Nationality Clean"] = (
            data[nationality_column]
            .apply(normalise_country)
        )

        data["Nationality Market"] = (
            data["Nationality Clean"]
            .apply(classify_market)
        )
    else:
        data["Nationality Clean"] = pd.NA
        data["Nationality Market"] = pd.NA

    if group_corporate_column:
        parsed_channel = data[
            group_corporate_column
        ].apply(parse_group_corporate)

        data["Registration Type"] = parsed_channel.str[0]
        data["Corporate Group"] = parsed_channel.str[1]
        data["Complimentary Programme"] = (
            parsed_channel.str[2]
        )
    else:
        data["Registration Type"] = "Public"
        data["Corporate Group"] = pd.NA
        data["Complimentary Programme"] = pd.NA

    # Add-on purchase flags: each source column is a 0/1 purchase
    # indicator. Values are coerced to numeric and clipped to 0/1
    # so stray text or blanks do not corrupt the daily counts.
    addon_columns = addon_columns or {}

    for addon_label, source_column in addon_columns.items():
        flag_values = pd.to_numeric(
            data[source_column],
            errors="coerce",
        ).fillna(0)

        data[f"Addon {addon_label}"] = (
            flag_values.gt(0).astype(int)
        )

    # A possible duplicate is a row where all supplied analytical
    # fields have identical values.
    duplicate_columns = [
        "Registration Date Time",
        "Grouped Category",
    ]

    if age_column:
        duplicate_columns.append("Current Age Clean")

    if gender_column:
        duplicate_columns.append("Gender Clean")

    if country_column:
        duplicate_columns.append("Country Clean")

    data["Possible Duplicate"] = data.duplicated(
        subset=duplicate_columns,
        keep=False,
    )

    return data


# =========================================================
# FILTERING
# =========================================================

def apply_filters(
    data,
    selected_date_range,
    selected_categories,
    selected_genders,
    selected_age_groups,
    selected_countries,
    selected_markets,
):
    """
    Apply sidebar selections to the prepared data.
    """
    filtered = data.copy()

    if selected_date_range and len(selected_date_range) == 2:
        start_date = pd.Timestamp(selected_date_range[0])
        end_date = pd.Timestamp(selected_date_range[1])

        filtered = filtered[
            filtered["Registration Date Only"].between(
                start_date,
                end_date,
                inclusive="both",
            )
        ]

    if selected_categories:
        filtered = filtered[
            filtered["Grouped Category"].isin(
                selected_categories
            )
        ]

    if selected_genders:
        filtered = filtered[
            filtered["Gender Clean"].isin(
                selected_genders
            )
        ]

    if selected_age_groups:
        filtered = filtered[
            filtered["Age Group"].isin(
                selected_age_groups
            )
        ]

    if selected_countries:
        filtered = filtered[
            filtered["Country Clean"].isin(
                selected_countries
            )
        ]

    if selected_markets:
        filtered = filtered[
            filtered["Market"].isin(
                selected_markets
            )
        ]

    return filtered


# =========================================================
# SUMMARY TABLES
# =========================================================

def create_daily_registration_summary(data):
    """
    Produce daily total registrations.
    """
    valid = data.dropna(
        subset=["Registration Date Only"]
    ).copy()

    return (
        valid.groupby("Registration Date Only")
        .size()
        .reset_index(name="Daily Registrations")
        .rename(
            columns={
                "Registration Date Only": "Date",
            }
        )
        .sort_values("Date")
        .reset_index(drop=True)
    )


def create_category_summary(data):
    """
    Produce category totals and shares.
    """
    valid = data[
        data["Grouped Category"].ne("Unmapped")
    ].copy()

    summary = (
        valid.groupby(
            "Grouped Category",
            observed=True,
        )
        .size()
        .reindex(CATEGORY_ORDER, fill_value=0)
        .rename("Registrations")
        .reset_index()
        .rename(
            columns={
                "Grouped Category": "Category",
            }
        )
    )

    total = summary["Registrations"].sum()

    if total > 0:
        summary["Share"] = (
            summary["Registrations"] / total * 100
        )
    else:
        summary["Share"] = 0.0

    return summary


def create_category_pace_table(data):
    """
    Compare the latest seven days with the preceding seven days.
    """
    valid = data[
        data["Registration Date Only"].notna()
        & data["Grouped Category"].ne("Unmapped")
    ].copy()

    if valid.empty:
        return pd.DataFrame(
            columns=[
                "Category",
                "Total Registrations",
                "Latest 7 Days",
                "Previous 7 Days",
                "Change",
                "Change %",
            ]
        )

    latest_date = valid["Registration Date Only"].max()

    latest_start = latest_date - pd.Timedelta(days=6)
    previous_end = latest_start - pd.Timedelta(days=1)
    previous_start = previous_end - pd.Timedelta(days=6)

    rows = []

    for category in CATEGORY_ORDER:
        category_data = valid[
            valid["Grouped Category"].eq(category)
        ]

        total = len(category_data)

        latest_count = int(
            category_data[
                category_data["Registration Date Only"].between(
                    latest_start,
                    latest_date,
                    inclusive="both",
                )
            ].shape[0]
        )

        previous_count = int(
            category_data[
                category_data["Registration Date Only"].between(
                    previous_start,
                    previous_end,
                    inclusive="both",
                )
            ].shape[0]
        )

        change = latest_count - previous_count

        if previous_count > 0:
            change_percentage = (
                change / previous_count * 100
            )
        elif latest_count > 0:
            change_percentage = 100.0
        else:
            change_percentage = 0.0

        rows.append(
            {
                "Category": category,
                "Total Registrations": total,
                "Latest 7 Days": latest_count,
                "Previous 7 Days": previous_count,
                "Change": change,
                "Change %": change_percentage,
            }
        )

    return pd.DataFrame(rows)


def create_daily_category_pivot(data):
    """
    Create daily registration counts with summarised categories as rows.
    """
    valid = data[
        data["Registration Date Only"].notna()
        & data["Grouped Category"].ne("Unmapped")
    ].copy()

    if valid.empty:
        return pd.DataFrame(index=CATEGORY_ORDER)

    daily_long = (
        valid.groupby(
            [
                "Registration Date Only",
                "Grouped Category",
            ],
            observed=True,
        )
        .size()
        .reset_index(name="Registrations")
    )

    pivot = (
        daily_long.pivot(
            index="Grouped Category",
            columns="Registration Date Only",
            values="Registrations",
        )
        .reindex(CATEGORY_ORDER)
        .fillna(0)
        .astype(int)
    )

    pivot = pivot.reindex(
        sorted(pivot.columns),
        axis=1,
    )

    pivot.index.name = "Category"

    return pivot


def create_tracker_table(data):
    """
    Create the cumulative category tracker and total rows.
    """
    daily_pivot = create_daily_category_pivot(data)

    if daily_pivot.empty:
        return daily_pivot

    cumulative_pivot = daily_pivot.cumsum(axis=1)

    tracker = cumulative_pivot.copy()

    tracker.loc["Cumulative"] = (
        cumulative_pivot.sum(axis=0)
    )

    tracker.loc["Daily Registrations"] = (
        daily_pivot.sum(axis=0)
    )

    tracker.index.name = "Category"

    return tracker


def create_channel_tracker_table(
    data, registration_type, mode="cumulative"
):
    """
    Category tracker restricted to one registration channel
    (Public/"Retail", Group Registration, or Complimentary).

    mode="cumulative" (default): category rows are the running
    total up to each date, with a "Cumulative" total row and a
    "Daily Registrations" row showing that day's new count for
    reference.

    mode="daily": category rows are that day's new registrations
    only (not accumulated), with a "Daily Total" row.
    """
    filtered = data[
        data["Registration Type"].eq(registration_type)
    ]

    daily_pivot = create_daily_category_pivot(filtered)

    if daily_pivot.empty:
        return daily_pivot

    if mode == "daily":
        tracker = daily_pivot.copy()

        tracker.loc["Daily Total"] = (
            daily_pivot.sum(axis=0)
        )
    else:
        cumulative_pivot = daily_pivot.cumsum(axis=1)

        tracker = cumulative_pivot.copy()

        tracker.loc["Cumulative"] = (
            cumulative_pivot.sum(axis=0)
        )

        tracker.loc["Daily Registrations"] = (
            daily_pivot.sum(axis=0)
        )

    tracker.index.name = "Category"

    return tracker


def create_all_channel_trackers(data, mode="cumulative"):
    """
    Build the three channel-specific trackers in one call: Retail
    (Public), Corporate (Group Registration) and Complimentary.
    Returns a dict keyed by display label. See
    create_channel_tracker_table for what mode controls.
    """
    return {
        "Retail (excl. corporate & complimentary)": (
            create_channel_tracker_table(
                data, "Public", mode=mode
            )
        ),
        "Corporate": create_channel_tracker_table(
            data, "Group Registration", mode=mode
        ),
        "Complimentary": create_channel_tracker_table(
            data, "Complimentary", mode=mode
        ),
    }


# =========================================================
# CORPORATE AND COMPLIMENTARY CHANNEL TABLES
# =========================================================

def generate_weekly_monday_snapshots(min_date, max_date):
    """
    Build the sequence of Monday as-of dates covering the
    campaign, matching the actual reporting cadence: every
    Tuesday, the total is pulled as of end of the previous day
    (Monday), inclusive.

    The first Monday is the first Monday on or after min_date.
    The last Monday is the most recent Monday on or before
    max_date — the snapshot is only as current as the data
    actually collected, so a still-in-progress week is not
    given its own (misleadingly low) column until the following
    Monday has been reached.
    """
    if pd.isna(min_date) or pd.isna(max_date):
        return []

    first_offset = (7 - min_date.weekday()) % 7
    first_monday = min_date + pd.Timedelta(days=first_offset)

    last_monday = max_date - pd.Timedelta(
        days=max_date.weekday()
    )

    if last_monday < first_monday:
        return []

    return list(
        pd.date_range(first_monday, last_monday, freq="7D")
    )


def build_weekly_snapshot_pivot(valid, group_keys, value_column, mondays):
    """
    For each group (e.g. race category, or category + market),
    return the cumulative sum of value_column as of end of each
    Monday snapshot date, inclusive.
    """
    daily = (
        valid.groupby(
            group_keys + ["Registration Date Only"],
            observed=True,
        )[value_column]
        .sum()
        .unstack(fill_value=0)
    )

    if daily.empty:
        return daily

    range_start = min(
        daily.columns.min(),
        mondays[0],
    )

    full_range = pd.date_range(
        range_start,
        mondays[-1],
        freq="D",
    )

    daily = daily.reindex(
        columns=full_range,
        fill_value=0,
    )

    cumulative = daily.cumsum(axis=1)

    result = cumulative[mondays]

    result.columns = [
        monday.strftime("%d %b %Y") for monday in mondays
    ]

    return result


def create_addon_market_category_trackers(data, addon_labels):
    """
    Weekly cumulative purchases per add-on, broken down by race
    category and Local/International market, mirroring the
    operational weekly purchase-tracking sheet: a Category x
    Market breakdown with a TOTAL row, plus a Summary table
    rolling Local and International back together per category.

    Matches the actual pull cadence: every Tuesday, the total is
    pulled as of end of the previous day (Monday), inclusive. So
    the column labelled with a given Monday's date includes every
    registration up to and including that Monday.

    Returns a dict of
    {add-on label: (breakdown DataFrame, summary DataFrame)}.
    Country values classified as Unknown are folded into
    International so every row still foots to the same total as
    the overall add-on tracker.
    """
    trackers = {}

    if not addon_labels:
        return trackers

    valid = data[
        data["Registration Date Only"].notna()
        & data["Grouped Category"].ne("Unmapped")
    ].copy()

    if valid.empty:
        return trackers

    mondays = generate_weekly_monday_snapshots(
        valid["Registration Date Only"].min(),
        valid["Registration Date Only"].max(),
    )

    if not mondays:
        return trackers

    valid["Market Label"] = valid["Market"].map(
        lambda market: (
            "Local" if market == "Singapore" else "International"
        )
    )

    market_row_index = pd.MultiIndex.from_product(
        [CATEGORY_ORDER, ["Local", "International"]],
        names=["Grouped Category", "Market Label"],
    )

    for addon_label in addon_labels:
        flag_column = f"Addon {addon_label}"

        breakdown_raw = build_weekly_snapshot_pivot(
            valid,
            ["Grouped Category", "Market Label"],
            flag_column,
            mondays,
        )

        if breakdown_raw.empty:
            breakdown = pd.DataFrame(
                0,
                index=market_row_index,
                columns=[
                    monday.strftime("%d %b %Y")
                    for monday in mondays
                ],
            )
        else:
            breakdown = breakdown_raw.reindex(
                market_row_index,
                fill_value=0,
            )

        breakdown.index = [
            f"{category} - {market_label}"
            for category, market_label in breakdown.index
        ]

        breakdown.loc["TOTAL"] = breakdown.sum(axis=0)

        breakdown.index.name = "Category"

        summary_raw = build_weekly_snapshot_pivot(
            valid,
            ["Grouped Category"],
            flag_column,
            mondays,
        )

        if summary_raw.empty:
            summary = pd.DataFrame(
                0,
                index=CATEGORY_ORDER,
                columns=[
                    monday.strftime("%d %b %Y")
                    for monday in mondays
                ],
            )
        else:
            summary = summary_raw.reindex(
                CATEGORY_ORDER,
                fill_value=0,
            )

        summary.index.name = "Category"

        trackers[addon_label] = (breakdown, summary)

    return trackers


# Short codes keep generated Excel sheet names (breakdown and
# "... Summary" variants) under the 31-character worksheet-name
# limit.
ADDON_SHEET_SHORT_NAMES = {
    "Gear Deposit Bag": "GearBag",
    "iTab": "iTab",
    "RunFlex": "RunFlex",
}


def create_registration_mix_table(data):
    """
    Registration counts split by channel — Public, Group
    Registration and Complimentary — in wide format: each channel
    is a row, each date is a column, with a Daily Total row and a
    Cumulative Total row.
    """
    valid = data[
        data["Registration Date Only"].notna()
    ]

    if valid.empty:
        return pd.DataFrame()

    mix = pd.pivot_table(
        valid,
        index="Registration Type",
        columns="Registration Date Only",
        aggfunc="size",
        fill_value=0,
    )

    mix = mix[sorted(mix.columns)]

    ordered_types = [
        channel
        for channel in REGISTRATION_TYPE_ORDER
        if channel in mix.index
    ] + [
        channel
        for channel in mix.index
        if channel not in REGISTRATION_TYPE_ORDER
    ]

    mix = mix.reindex(ordered_types, fill_value=0)

    mix.loc["Daily Total"] = mix.sum(axis=0)

    mix.loc["Cumulative Total"] = (
        mix.loc["Daily Total"].cumsum()
    )

    mix.columns = [
        date.strftime("%d %b %Y") for date in mix.columns
    ]

    mix.index.name = "Channel"

    return mix


def create_addon_summary_table(data, addon_labels):
    """
    Daily add-on purchase counts in wide format: each add-on is a
    row, each date is a column, with a Total column (grand total
    per add-on) and a Total Add-Ons Purchased row (all add-ons
    combined, per day and overall).
    """
    if not addon_labels:
        return pd.DataFrame()

    valid = data[
        data["Registration Date Only"].notna()
    ]

    if valid.empty:
        return pd.DataFrame()

    flag_columns = [
        f"Addon {label}" for label in addon_labels
    ]

    daily = (
        valid.groupby("Registration Date Only")[
            flag_columns
        ]
        .sum()
        .sort_index()
    )

    wide = daily.T
    wide.index = addon_labels

    wide["Total"] = daily.sum(axis=0).to_numpy()

    date_columns = [
        date.strftime("%d %b %Y") for date in daily.index
    ]

    wide.columns = date_columns + ["Total"]

    wide.loc["Total Add-Ons Purchased"] = wide.sum(axis=0)

    wide.index.name = "Add-On"

    return wide


def create_complimentary_daily_table(data):
    """
    Daily complimentary sign-ups broken down by programme (the
    identity after the COMPLIMENTARY_ prefix), with a daily total
    column and programme totals row.
    """
    complimentary = data[
        data["Registration Date Only"].notna()
        & data["Registration Type"].eq("Complimentary")
    ]

    if complimentary.empty:
        return pd.DataFrame()

    daily = (
        complimentary.pivot_table(
            index="Registration Date Only",
            columns="Complimentary Programme",
            aggfunc="size",
            fill_value=0,
        )
        .sort_index()
    )

    # Order programmes by overall size so the biggest allocations
    # sit on the left.
    daily = daily[
        daily.sum().sort_values(ascending=False).index
    ]

    daily["Daily Total"] = daily.sum(axis=1)

    daily.index = daily.index.strftime("%d %b %Y")
    daily.index.name = "Date"

    daily.loc["Programme Total"] = daily.sum(axis=0)

    return daily


# =========================================================
# EXECUTIVE SNAPSHOT TABLES
# =========================================================

def create_market_split_table(data, market_column, label):
    """
    Local ("<label>") vs International split with counts and
    percentages, plus a Grand Total row. Used for both the
    Country-based and, where available, the Nationality-based
    split on the Executive Snapshot.
    """
    valid = data[market_column].dropna()
    valid = valid[valid.ne("Unknown")]

    if valid.empty:
        return pd.DataFrame()

    counts = valid.value_counts()

    local_count = int(counts.get("Singapore", 0))
    international_count = int(counts.get("International", 0))
    total = local_count + international_count

    if total == 0:
        return pd.DataFrame()

    table = pd.DataFrame(
        {
            label: [local_count, international_count, total],
            f"{label} %": [
                f"{local_count / total * 100:.0f}%",
                f"{international_count / total * 100:.0f}%",
                "100%",
            ],
        },
        index=["Singapore", "Non-Singapore", "Grand Total"],
    )

    table.index.name = "Market"

    return table


def create_age_gender_table(data):
    """
    Age band x Gender counts and percentages, with a Grand Total
    row, matching the Executive Snapshot's Age & Gender Splits.
    """
    valid = data[
        data["Age Group"].ne("Unknown / Invalid")
    ]

    if valid.empty:
        return pd.DataFrame()

    counts = pd.crosstab(
        valid["Age Group"], valid["Gender Clean"]
    )

    for gender in ["Male", "Female"]:
        if gender not in counts.columns:
            counts[gender] = 0

    counts = counts[["Male", "Female"]].reindex(
        [g for g in AGE_GROUP_ORDER if g != "Unknown / Invalid"],
        fill_value=0,
    )

    counts["Total"] = counts["Male"] + counts["Female"]

    grand_total = int(counts["Total"].sum())

    table = counts.copy()

    for gender in ["Male", "Female", "Total"]:
        table[f"{gender} %"] = (
            counts[gender] / grand_total * 100
        ).map(lambda value: f"{value:.0f}%")

    table = table[
        ["Male", "Male %", "Female", "Female %", "Total", "Total %"]
    ]

    table.loc["Grand Total"] = [
        int(counts["Male"].sum()),
        "100%" if counts["Male"].sum() else "0%",
        int(counts["Female"].sum()),
        "100%" if counts["Female"].sum() else "0%",
        grand_total,
        "100%",
    ]

    # Grand Total row's Male/Female % should be their own share of
    # the grand total, not a flat 100%.
    table.loc["Grand Total", "Male %"] = (
        f"{counts['Male'].sum() / grand_total * 100:.0f}%"
    )
    table.loc["Grand Total", "Female %"] = (
        f"{counts['Female'].sum() / grand_total * 100:.0f}%"
    )

    table.index.name = "Age Group"

    return table


def create_category_gender_table(data):
    """
    Race category x Gender counts and percentages, with a Grand
    Total row, matching the Executive Snapshot's Category &
    Gender Splits.
    """
    valid = data[
        data["Grouped Category"].ne("Unmapped")
    ]

    if valid.empty:
        return pd.DataFrame()

    counts = pd.crosstab(
        valid["Grouped Category"], valid["Gender Clean"]
    )

    for gender in ["Male", "Female"]:
        if gender not in counts.columns:
            counts[gender] = 0

    counts = counts[["Male", "Female"]].reindex(
        CATEGORY_ORDER, fill_value=0
    )

    counts["Total"] = counts["Male"] + counts["Female"]

    grand_total = int(counts["Total"].sum())

    table = counts.copy()

    for gender in ["Male", "Female", "Total"]:
        table[f"{gender} %"] = (
            counts[gender] / grand_total * 100
        ).map(lambda value: f"{value:.0f}%")

    table = table[
        ["Male", "Male %", "Female", "Female %", "Total", "Total %"]
    ]

    table.loc["Grand Total"] = [
        int(counts["Male"].sum()),
        f"{counts['Male'].sum() / grand_total * 100:.0f}%",
        int(counts["Female"].sum()),
        f"{counts['Female'].sum() / grand_total * 100:.0f}%",
        grand_total,
        "100%",
    ]

    table.index.name = "Category"

    return table


def create_age_category_gender_grid(data):
    """
    The combined Age Range x Category x Gender grid: for each age
    band, every race category's Male/Female/Total counts and
    percentages of the campaign grand total, followed by an
    "<Age Band> Total" row. Flat layout (age band repeated per
    row) so it renders as one scrollable table.
    """
    valid = data[
        data["Age Group"].ne("Unknown / Invalid")
        & data["Grouped Category"].ne("Unmapped")
    ]

    if valid.empty:
        return pd.DataFrame()

    grand_total = len(valid)

    age_bands = [
        band
        for band in AGE_GROUP_ORDER
        if band != "Unknown / Invalid"
    ]

    rows = []

    for age_band in age_bands:
        band_data = valid[valid["Age Group"].eq(age_band)]

        counts = pd.crosstab(
            band_data["Grouped Category"],
            band_data["Gender Clean"],
        )

        for gender in ["Male", "Female"]:
            if gender not in counts.columns:
                counts[gender] = 0

        counts = counts.reindex(CATEGORY_ORDER, fill_value=0)

        for category in CATEGORY_ORDER:
            male = int(counts.loc[category, "Male"])
            female = int(counts.loc[category, "Female"])
            total = male + female

            if total == 0:
                continue

            rows.append(
                {
                    "Age Range": age_band,
                    "Category": category,
                    "Male": male,
                    "Male %": f"{male / grand_total * 100:.0f}%",
                    "Female": female,
                    "Female %": (
                        f"{female / grand_total * 100:.0f}%"
                    ),
                    "Total": total,
                    "Total %": (
                        f"{total / grand_total * 100:.0f}%"
                    ),
                }
            )

        band_male = int(counts["Male"].sum())
        band_female = int(counts["Female"].sum())
        band_total = band_male + band_female

        rows.append(
            {
                "Age Range": age_band,
                "Category": f"{age_band} Total",
                "Male": band_male,
                "Male %": (
                    f"{band_male / grand_total * 100:.0f}%"
                ),
                "Female": band_female,
                "Female %": (
                    f"{band_female / grand_total * 100:.0f}%"
                ),
                "Total": band_total,
                "Total %": (
                    f"{band_total / grand_total * 100:.0f}%"
                ),
            }
        )

    return pd.DataFrame(rows)


# Segment order for the capacity/channel breakdown table, matching
# the operational capacity-tracking sheet. Elite, National
# Championship and Score Collab are placeholder rows: as of this
# build there is no reliable rule in the registration export to
# identify them (no promo code or field is dedicated to them yet),
# so they are always shown at 0 until that identifying rule exists.
CAPACITY_SEGMENT_ORDER = [
    "Standard Chartered Staff Entry",
    "Corporate Registration",
    "Complimentary Entry",
    "Elite",
    "National Championship",
    "Local",
    "International",
    "Score Collab",
    "Unsuccessful Medic Entry",
    "Transfer Entry",
]

# Always-zero placeholder segments (see note above).
CAPACITY_PLACEHOLDER_SEGMENTS = [
    "Elite",
    "National Championship",
    "Score Collab",
]

# Capacity reporting groups: each maps to one "main" race category
# for its channel breakdown, plus an optional nested category
# (BYD Marathon Crew Challenge is its own registration category in
# the data, but is reported as a single row nested under Marathon
# for capacity planning, since crew runners are budgeted against
# the Marathon's capacity).
CAPACITY_GROUPS = {
    "Marathon": {
        "main": "BYD Marathon",
        "nested": [("Crew Challenge", "BYD Marathon Crew Challenge")],
    },
    "Half Marathon": {"main": "adidas Half Marathon", "nested": []},
    "10km": {"main": "Standard Chartered 10km", "nested": []},
    "5km": {"main": "5km", "nested": []},
    "Kids Dash Competitive 1.6KM": {
        "main": "Kids Dash Competitive 1.6KM",
        "nested": [],
    },
    "Kids Dash Non-Competitive 1.6KM": {
        "main": "Kids Dash Non-Competitive 1.6KM",
        "nested": [],
    },
    "Kids Dash Non-Competitive 600m": {
        "main": "Kids Dash Non-Competitive 600m",
        "nested": [],
    },
}


def classify_capacity_segment(row):
    """
    Assign one registration row to a capacity-table segment.
    Order matters: the more specific checks (unsuccessful medic,
    SC staff) must run before the generic corporate/public
    fallbacks, or they would be absorbed into "Corporate
    Registration" or "Local"/"International".

    The split between "Corporate Registration" and a possible
    future "Standard Chartered Corporate Buy" row is intentionally
    on hold — all non-staff Group Registration rows currently roll
    up into "Corporate Registration".
    """
    corporate_name = row.get("Corporate Group") or ""
    group_corporate_raw = row.get("_group_corporate_raw") or ""
    category_raw = row.get("_category_raw") or ""

    if "UNSFUL_MEDIC" in str(group_corporate_raw).upper():
        return "Unsuccessful Medic Entry"

    if "Transfer Entry" in str(category_raw):
        return "Transfer Entry"

    if row["Registration Type"] == "Group Registration":
        if "Standard_Chartered_Staff" in str(group_corporate_raw):
            return "Standard Chartered Staff Entry"

        return "Corporate Registration"

    if row["Registration Type"] == "Complimentary":
        return "Complimentary Entry"

    if row["Market"] == "Singapore":
        return "Local"

    return "International"


def create_capacity_segment_table(
    data,
    group_corporate_column=None,
    category_column=None,
):
    """
    Actual registered/sold counts per capacity segment, by race
    category group (Marathon, Half Marathon, 10km, 5km, Kids
    Dash), with a TOTAL row and TOTAL column. Slots/Remaining
    columns are intentionally not produced here — those are
    planning targets maintained outside the registration system,
    not something the export can supply.
    """
    valid = data[
        data["Grouped Category"].ne("Unmapped")
    ].copy()

    if valid.empty:
        return pd.DataFrame()

    valid["_group_corporate_raw"] = (
        data[group_corporate_column]
        if group_corporate_column
        else ""
    )

    valid["_category_raw"] = (
        data[category_column] if category_column else ""
    )

    valid["_segment"] = valid.apply(
        classify_capacity_segment, axis=1
    )

    table = pd.DataFrame(
        0,
        index=CAPACITY_SEGMENT_ORDER,
        columns=list(CAPACITY_GROUPS.keys()),
    )

    for group_name, group_definition in CAPACITY_GROUPS.items():
        main_category = group_definition["main"]

        group_rows = valid[
            valid["Grouped Category"].eq(main_category)
        ]

        segment_counts = group_rows["_segment"].value_counts()

        for segment in CAPACITY_SEGMENT_ORDER:
            if segment in CAPACITY_PLACEHOLDER_SEGMENTS:
                continue

            table.loc[segment, group_name] = int(
                segment_counts.get(segment, 0)
            )

        for nested_label, nested_category in group_definition[
            "nested"
        ]:
            nested_count = int(
                valid["Grouped Category"]
                .eq(nested_category)
                .sum()
            )

            if nested_label not in table.index:
                table.loc[nested_label] = 0

            table.loc[nested_label, group_name] = nested_count

    table["TOTAL"] = table.sum(axis=1)
    table.loc["TOTAL"] = table.sum(axis=0)

    table.index.name = "Segment"

    return table


def create_corporate_category_table(data):
    """
    For each corporate group, the count of registrations in each
    race category. Returns the numeric crosstab (for charting)
    and a display version where every cell reads "count (share of
    the company's registrations)".
    """
    corporate = data[
        data["Registration Date Only"].notna()
        & data["Registration Type"].eq(
            "Group Registration"
        )
        & data["Corporate Group"].notna()
    ]

    if corporate.empty:
        return pd.DataFrame(), pd.DataFrame()

    counts = pd.crosstab(
        corporate["Corporate Group"],
        corporate["Grouped Category"],
    )

    ordered_categories = [
        category
        for category in CATEGORY_ORDER
        if category in counts.columns
    ] + [
        category
        for category in counts.columns
        if category not in CATEGORY_ORDER
    ]

    counts = counts[ordered_categories]

    counts["Total"] = counts.sum(axis=1)

    counts = counts.sort_values(
        "Total",
        ascending=False,
    )

    counts.index.name = "Corporate Group"

    display = counts.copy().astype(object)

    for company in counts.index:
        company_total = counts.loc[company, "Total"]

        for category in ordered_categories:
            cell_count = counts.loc[company, category]

            if cell_count > 0 and company_total > 0:
                share = cell_count / company_total * 100

                display.loc[company, category] = (
                    f"{cell_count:,} ({share:.0f}%)"
                )
            else:
                display.loc[company, category] = "—"

        display.loc[company, "Total"] = (
            f"{company_total:,}"
        )

    return counts, display


def create_complimentary_category_table(data):
    """
    For each complimentary programme, the count of registrations
    in each race category. Same shape as
    create_corporate_category_table, keyed on Complimentary
    Programme instead of Corporate Group. Returns the numeric
    crosstab (for charting) and a display version where every
    cell reads "count (share of the programme's registrations)".
    """
    complimentary = data[
        data["Registration Date Only"].notna()
        & data["Registration Type"].eq("Complimentary")
        & data["Complimentary Programme"].notna()
    ]

    if complimentary.empty:
        return pd.DataFrame(), pd.DataFrame()

    counts = pd.crosstab(
        complimentary["Complimentary Programme"],
        complimentary["Grouped Category"],
    )

    ordered_categories = [
        category
        for category in CATEGORY_ORDER
        if category in counts.columns
    ] + [
        category
        for category in counts.columns
        if category not in CATEGORY_ORDER
    ]

    counts = counts[ordered_categories]

    counts["Total"] = counts.sum(axis=1)

    counts = counts.sort_values(
        "Total",
        ascending=False,
    )

    counts.index.name = "Complimentary Programme"

    display = counts.copy().astype(object)

    for programme in counts.index:
        programme_total = counts.loc[programme, "Total"]

        for category in ordered_categories:
            cell_count = counts.loc[programme, category]

            if cell_count > 0 and programme_total > 0:
                share = cell_count / programme_total * 100

                display.loc[programme, category] = (
                    f"{cell_count:,} ({share:.0f}%)"
                )
            else:
                display.loc[programme, category] = "—"

        display.loc[programme, "Total"] = (
            f"{programme_total:,}"
        )

    return counts, display


def create_age_summary(data):
    """
    Produce participant count by age group.
    """
    counts = (
        data["Age Group"]
        .value_counts()
        .reindex(AGE_GROUP_ORDER, fill_value=0)
        .rename_axis("Age Group")
        .reset_index(name="Participants")
    )

    return counts


def create_gender_summary(data):
    """
    Produce participant count by gender.
    """
    return (
        data["Gender Clean"]
        .value_counts()
        .rename_axis("Gender")
        .reset_index(name="Participants")
    )


def create_country_summary(data):
    """
    Produce participant count by country.
    """
    return (
        data["Country Clean"]
        .value_counts()
        .rename_axis("Country")
        .reset_index(name="Participants")
    )


def create_data_quality_summary(
    data,
    age_available,
    gender_available,
    country_available,
):
    """
    Produce a concise data-quality report.
    """
    total_rows = len(data)

    invalid_dates = int(
        data["Registration Date Time"]
        .isna()
        .sum()
    )

    unmapped_categories = int(
        data["Grouped Category"]
        .eq("Unmapped")
        .sum()
    )

    if age_available:
        invalid_age = int(
            (
                data["Current Age Clean"].isna()
                | data["Current Age Clean"].lt(0)
                | data["Current Age Clean"].gt(110)
            ).sum()
        )
    else:
        invalid_age = total_rows

    if gender_available:
        unknown_gender = int(
            data["Gender Clean"]
            .eq("Unknown")
            .sum()
        )
    else:
        unknown_gender = total_rows

    if country_available:
        unknown_country = int(
            data["Country Clean"]
            .eq("Unknown")
            .sum()
        )
    else:
        unknown_country = total_rows

    possible_duplicates = int(
        data["Possible Duplicate"]
        .sum()
    )

    return pd.DataFrame(
        [
            {
                "Data Quality Check": "Total source rows",
                "Affected Records": total_rows,
            },
            {
                "Data Quality Check": "Invalid registration dates",
                "Affected Records": invalid_dates,
            },
            {
                "Data Quality Check": "Unmapped categories",
                "Affected Records": unmapped_categories,
            },
            {
                "Data Quality Check": "Missing or invalid ages",
                "Affected Records": invalid_age,
            },
            {
                "Data Quality Check": "Unknown gender",
                "Affected Records": unknown_gender,
            },
            {
                "Data Quality Check": "Unknown country",
                "Affected Records": unknown_country,
            },
            {
                "Data Quality Check": "Possible duplicate rows",
                "Affected Records": possible_duplicates,
            },
        ]
    )


def parse_group_corporate(value):
    """
    Split the mixed Group/Corporate Name field into its two
    identities.

    Returns (registration type, corporate group, complimentary
    programme). Empty cells are Public registrations. For group
    registrations, batch numbers and coordinator names after the
    company (e.g. "SCB Batch 1 - Tan, Hui Hoon") are stripped so
    every batch rolls up to one company, and known label variants
    are folded together via COMPANY_ALIASES.
    """
    if value is None or (
        isinstance(value, float) and pd.isna(value)
    ):
        return ("Public", None, None)

    text = str(value).strip()

    if not text or text.lower() in ("nan", "none"):
        return ("Public", None, None)

    upper_text = text.upper()

    if upper_text.startswith(COMPLIMENTARY_PREFIX):
        programme = text[
            len(COMPLIMENTARY_PREFIX):
        ].strip("_- ")

        programme = (
            programme.replace("_", " ").strip()
            or "Unspecified"
        )

        return ("Complimentary", None, programme)

    if upper_text.startswith(GROUP_REGISTRATION_PREFIX):
        remainder = text[
            len(GROUP_REGISTRATION_PREFIX):
        ].strip("_- ")

        # Drop batch suffixes and the coordinator names that
        # follow them: "SCB Batch 1 - Tan, Hui Hoon" -> "SCB".
        remainder = re.split(
            r"(?i)\s+Batch\s+\d+",
            remainder,
        )[0]

        remainder = remainder.split(" - ")[0]

        company = (
            remainder.strip()
            .strip("_")
            .replace("_", " ")
            .strip()
            or "Unspecified"
        )

        company = COMPANY_ALIASES.get(company, company)

        return ("Group Registration", company, None)

    # A non-empty value with neither known prefix: keep it visible
    # rather than silently absorbing it into a known channel.
    return ("Other", None, None)


# =========================================================
# TARGET PACING AND PROJECTION
# =========================================================

def calculate_run_rate(daily_counts):
    """
    Average daily registrations over the last seven complete days.

    Returns the run rate and the last date treated as complete.
    """
    if daily_counts.empty:
        return 0.0, None

    counts = daily_counts.sort_values("Date").copy()

    if EXCLUDE_LATEST_DAY_FROM_PACE and len(counts) > 1:
        counts = counts.iloc[:-1]

    window = counts.tail(7)

    run_rate = float(
        window["Daily Registrations"].mean()
    )

    return run_rate, counts["Date"].max()


def build_complete_daily_counts(daily_summary):
    """
    Reindex the active-day summary over the full calendar range so
    zero-registration days count as zeros — otherwise quiet days
    silently inflate every average and forecast.
    """
    if daily_summary.empty:
        return pd.DataFrame(
            columns=["Date", "Daily Registrations"]
        )

    counts = daily_summary.set_index("Date")[
        "Daily Registrations"
    ]

    full_range = pd.date_range(
        counts.index.min(),
        counts.index.max(),
        freq="D",
    )

    return (
        counts.reindex(full_range, fill_value=0)
        .rename_axis("Date")
        .reset_index(name="Daily Registrations")
    )


def calculate_weekday_profile(complete_counts):
    """
    Mean registrations per weekday over the most recent 28 complete
    days, excluding the launch-day surge. Returns the per-weekday
    means, the residual standard deviation (forecast noise), and
    the number of days the statistics rest on.
    """
    counts = complete_counts.copy()

    if EXCLUDE_LATEST_DAY_FROM_PACE and len(counts) > 1:
        counts = counts.iloc[:-1]

    window = counts.tail(28)

    stats_window = window[
        window["Date"].ne(REGISTRATION_OPEN_DATE)
    ].copy()

    if stats_window.empty:
        return {}, 0.0, 0

    overall_mean = float(
        stats_window["Daily Registrations"].mean()
    )

    stats_window["Weekday"] = stats_window[
        "Date"
    ].dt.day_name()

    weekday_means = (
        stats_window.groupby("Weekday")[
            "Daily Registrations"
        ]
        .mean()
        .to_dict()
    )

    # Any weekday not present in the window falls back to the
    # overall mean rather than zero.
    for weekday in WEEKDAY_ORDER:
        weekday_means.setdefault(weekday, overall_mean)

    residuals = stats_window[
        "Daily Registrations"
    ] - stats_window["Weekday"].map(weekday_means)

    residual_std = (
        float(residuals.std(ddof=1))
        if len(residuals) > 1
        else 0.0
    )

    return weekday_means, residual_std, len(stats_window)


def create_weekday_adjusted_projection(daily_summary):
    """
    Forecast cumulative registrations to the close date with a
    day-of-week adjusted model and a 95% confidence band.

    Each future day is forecast at its weekday's recent mean, so
    the projection no longer inherits whichever weekday mix the
    last few days happened to contain. The band widens with the
    square root of the days ahead, reflecting compounding
    uncertainty. With under 14 days of history the model degrades
    gracefully to a flat mean with a wider, less trustworthy band.
    """
    empty_result = pd.DataFrame(
        columns=[
            "Date",
            "Projected Cumulative",
            "Lower",
            "Upper",
        ]
    )

    if daily_summary.empty:
        return empty_result

    latest_date = daily_summary["Date"].max()

    if latest_date >= REGISTRATION_CLOSE_DATE:
        return empty_result

    complete_counts = build_complete_daily_counts(
        daily_summary
    )

    weekday_means, residual_std, window_size = (
        calculate_weekday_profile(complete_counts)
    )

    if window_size == 0:
        return empty_result

    current_total = int(
        daily_summary["Daily Registrations"].sum()
    )

    future_dates = pd.date_range(
        latest_date,
        REGISTRATION_CLOSE_DATE,
        freq="D",
    )

    cumulative = float(current_total)
    variance = 0.0

    rows = []

    for offset, future_date in enumerate(future_dates):
        if offset > 0:
            cumulative += weekday_means[
                future_date.day_name()
            ]
            variance += residual_std ** 2

        band = 1.96 * (variance ** 0.5)

        rows.append(
            {
                "Date": future_date,
                "Projected Cumulative": cumulative,
                "Lower": max(
                    cumulative - band,
                    float(current_total),
                ),
                "Upper": cumulative + band,
            }
        )

    return pd.DataFrame(rows)


def forecast_group_remaining(
    group_data,
    overall_start,
    overall_latest,
):
    """
    Point forecast of a target group's additional registrations
    between the data cut and the close date, using the group's own
    weekday profile. Groups with no recent activity correctly
    project near zero because quiet days are counted as zeros.
    """
    if overall_latest >= REGISTRATION_CLOSE_DATE:
        return 0.0

    group_daily = (
        group_data.groupby("Registration Date Only")
        .size()
        .reindex(
            pd.date_range(
                overall_start,
                overall_latest,
                freq="D",
            ),
            fill_value=0,
        )
        .rename_axis("Date")
        .reset_index(name="Daily Registrations")
    )

    weekday_means, _, window_size = (
        calculate_weekday_profile(group_daily)
    )

    if window_size == 0:
        return 0.0

    remaining_dates = pd.date_range(
        overall_latest + pd.Timedelta(days=1),
        REGISTRATION_CLOSE_DATE,
        freq="D",
    )

    return float(
        sum(
            weekday_means[future_date.day_name()]
            for future_date in remaining_dates
        )
    )


def calculate_target_progress(data):
    """
    Compare each target group's current registrations and projected
    close-date total against its management target.
    """
    valid = data[
        data["Registration Date Only"].notna()
        & data["Grouped Category"].ne("Unmapped")
    ].copy()

    if valid.empty:
        overall_start = None
        overall_latest = None
    else:
        overall_start = valid["Registration Date Only"].min()
        overall_latest = valid["Registration Date Only"].max()

    rows = []

    for group_name, categories in TARGET_GROUPS.items():
        target = CATEGORY_TARGETS.get(group_name)

        group_data = valid[
            valid["Grouped Category"].isin(categories)
        ]

        current = len(group_data)

        if overall_latest is None:
            projected = float(current)
        else:
            projected = current + forecast_group_remaining(
                group_data,
                overall_start,
                overall_latest,
            )

        if target:
            fill_percentage = current / target * 100

            if current >= target:
                status = "Target met"
            elif projected >= target:
                status = "On track"
            elif projected >= target * 0.9:
                status = "At risk"
            else:
                status = "Off track"
        else:
            fill_percentage = None
            status = "No target"

        rows.append(
            {
                "Target Group": group_name,
                "Target": target,
                "Current": current,
                "Fill %": fill_percentage,
                "Projected at Close": round(projected),
                "Status": status,
            }
        )

    return pd.DataFrame(rows)


def summarise_daily_pattern(daily_summary):
    """
    Describe the daily registration pattern with the launch-day
    surge separated out, so call-outs highlight genuinely
    informative days rather than the structural launch spike.
    """
    if daily_summary.empty:
        return None

    counts = daily_summary.sort_values("Date").copy()

    launch_rows = counts[
        counts["Date"].eq(REGISTRATION_OPEN_DATE)
    ]

    launch_count = (
        int(launch_rows["Daily Registrations"].iloc[0])
        if not launch_rows.empty
        else None
    )

    campaign_days = counts[
        counts["Date"].ne(REGISTRATION_OPEN_DATE)
    ]

    if campaign_days.empty:
        return {
            "launch_count": launch_count,
            "median": None,
            "best_date": None,
            "best_count": None,
            "above_median_last_14": None,
            "days_in_window": 0,
        }

    median_value = float(
        campaign_days["Daily Registrations"].median()
    )

    best_row = campaign_days.loc[
        campaign_days["Daily Registrations"].idxmax()
    ]

    recent_window = campaign_days.tail(14)

    above_median_recent = int(
        recent_window["Daily Registrations"]
        .gt(median_value)
        .sum()
    )

    return {
        "launch_count": launch_count,
        "median": median_value,
        "best_date": best_row["Date"],
        "best_count": int(best_row["Daily Registrations"]),
        "above_median_last_14": above_median_recent,
        "days_in_window": len(recent_window),
    }


def add_milestone_markers(figure):
    """
    Draw configured campaign milestones on a time-series figure.
    """
    for date_text, label in MILESTONES.items():
        milestone_date = pd.Timestamp(date_text)

        figure.add_vline(
            x=milestone_date,
            line_dash="dot",
            line_color=UNKNOWN_COLOR,
        )

        figure.add_annotation(
            x=milestone_date,
            y=1.02,
            yref="paper",
            text=label,
            showarrow=False,
            font={"size": 10, "color": UNKNOWN_COLOR},
        )

    return figure


# =========================================================
# KPI CALCULATIONS
# =========================================================

def calculate_registration_kpis(data):
    """
    Calculate management KPIs based on the selected data.
    """
    valid = data.dropna(
        subset=["Registration Date Only"]
    ).copy()

    if valid.empty:
        return {
            "total": 0,
            "latest_date": None,
            "latest_day": 0,
            "previous_day": 0,
            "same_weekday_prior": 0,
            "days_to_close": 0,
            "latest_7_days": 0,
            "previous_7_days": 0,
            "seven_day_average": 0.0,
            "week_change_percentage": 0.0,
            "countries": 0,
            "top_category": "Not available",
            "top_category_count": 0,
        }

    latest_date = valid["Registration Date Only"].max()

    latest_day = int(
        valid[
            valid["Registration Date Only"].eq(
                latest_date
            )
        ].shape[0]
    )

    previous_date = latest_date - pd.Timedelta(days=1)

    previous_day = int(
        valid[
            valid["Registration Date Only"].eq(
                previous_date
            )
        ].shape[0]
    )

    # Same weekday one week earlier removes weekday seasonality from
    # the day-on-day comparison (Mondays compare with Mondays).
    same_weekday_date = latest_date - pd.Timedelta(days=7)

    same_weekday_prior = int(
        valid[
            valid["Registration Date Only"].eq(
                same_weekday_date
            )
        ].shape[0]
    )

    days_to_close = max(
        (REGISTRATION_CLOSE_DATE - latest_date).days,
        0,
    )

    latest_week_start = latest_date - pd.Timedelta(days=6)

    latest_7_days = int(
        valid[
            valid["Registration Date Only"].between(
                latest_week_start,
                latest_date,
                inclusive="both",
            )
        ].shape[0]
    )

    previous_week_end = latest_week_start - pd.Timedelta(days=1)
    previous_week_start = previous_week_end - pd.Timedelta(days=6)

    previous_7_days = int(
        valid[
            valid["Registration Date Only"].between(
                previous_week_start,
                previous_week_end,
                inclusive="both",
            )
        ].shape[0]
    )

    seven_day_average = latest_7_days / 7

    if previous_7_days > 0:
        week_change_percentage = (
            (latest_7_days - previous_7_days)
            / previous_7_days
            * 100
        )
    elif latest_7_days > 0:
        week_change_percentage = 100.0
    else:
        week_change_percentage = 0.0

    country_values = valid[
        valid["Country Clean"].ne("Unknown")
    ]["Country Clean"]

    countries = int(country_values.nunique())

    category_counts = (
        valid[
            valid["Grouped Category"].ne("Unmapped")
        ]["Grouped Category"]
        .value_counts()
    )

    if category_counts.empty:
        top_category = "Not available"
        top_category_count = 0
    else:
        top_category = str(category_counts.index[0])
        top_category_count = int(category_counts.iloc[0])

    return {
        "total": len(valid),
        "latest_date": latest_date,
        "latest_day": latest_day,
        "previous_day": previous_day,
        "same_weekday_prior": same_weekday_prior,
        "days_to_close": days_to_close,
        "latest_7_days": latest_7_days,
        "previous_7_days": previous_7_days,
        "seven_day_average": seven_day_average,
        "week_change_percentage": week_change_percentage,
        "countries": countries,
        "top_category": top_category,
        "top_category_count": top_category_count,
    }


# =========================================================
# EXCEL FORMATTING
# =========================================================

def format_standard_worksheet(worksheet):
    """
    Apply common Excel formatting.
    """
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="4472C4",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column_cells in worksheet.columns:
        maximum_length = 0

        for cell in column_cells:
            if cell.value is not None:
                maximum_length = max(
                    maximum_length,
                    len(str(cell.value)),
                )

        column_letter = get_column_letter(
            column_cells[0].column
        )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(maximum_length + 2, 12),
            45,
        )


def format_tracker_worksheet(worksheet):
    """
    Apply formatting to the cumulative registration tracker.
    """
    worksheet.freeze_panes = "B2"

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="4472C4",
    )

    category_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    cumulative_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    )

    daily_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    bold_font = Font(bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.column_dimensions["A"].width = 42

    for column_number in range(
        2,
        worksheet.max_column + 1,
    ):
        worksheet.cell(
            row=1,
            column=column_number,
        ).number_format = "d-mmm"

        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = 12

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        label = worksheet.cell(
            row=row_number,
            column=1,
        ).value

        if label == "Cumulative":
            row_fill = cumulative_fill
            row_font = bold_font

        elif label == "Daily Registrations":
            row_fill = daily_fill
            row_font = bold_font

        else:
            row_fill = category_fill
            row_font = Font(bold=False)

        for cell in worksheet[row_number]:
            cell.fill = row_fill
            cell.font = row_font

            if cell.column > 1:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )


# =========================================================
# EXCEL EXPORT
# =========================================================

def create_archive_csv_bundle(tables):
    """
    Stack multiple tables into a single CSV for archiving: each
    table is preceded by a plain-text header row naming it, then
    its own header row and data, then a blank separator row
    before the next table. Tables that are None or empty are
    skipped. Any DataFrame index with a name (Category, Segment,
    Corporate Group, etc.) is written out as its own column so
    row labels are not lost.

    Because the tables have very different shapes and column
    counts (a handful of columns for a demographic split, 90+
    date columns for a daily tracker), this file will look
    visually uneven if opened directly in Excel — that unevenness
    is expected, and does not affect re-parsing the file later.
    """
    output = StringIO()

    for table_name, table in tables.items():
        if table is None or table.empty:
            continue

        output.write(f"{table_name}\n")

        export_table = table.copy()

        if not isinstance(export_table.index, pd.RangeIndex):
            if export_table.index.name is None:
                export_table.index.name = table_name

            export_table = export_table.reset_index()

        export_table.columns = [
            column.strftime("%d-%b-%Y")
            if isinstance(column, pd.Timestamp)
            else column
            for column in export_table.columns
        ]

        export_table.to_csv(output, index=False)

        output.write("\n")

    return output.getvalue().encode("utf-8-sig")


def create_excel_report(
    original_data,
    filtered_data,
    category_summary,
    category_pace,
    daily_summary,
    age_summary,
    gender_summary,
    country_summary,
    quality_summary,
    tracker_table,
    target_progress,
    registration_mix_table,
    corporate_category_display,
    addon_summary_table,
    addon_market_category_trackers,
    country_market_table,
    nationality_market_table,
    age_gender_table,
    category_gender_table,
    age_category_gender_grid,
    capacity_segment_table,
    channel_trackers,
    channel_trackers_daily,
    complimentary_category_display,
):
    """
    Generate a downloadable Excel workbook.
    """
    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        category_summary.to_excel(
            writer,
            sheet_name="Category Summary",
            index=False,
        )

        category_pace.to_excel(
            writer,
            sheet_name="Category Pace",
            index=False,
        )

        target_progress.to_excel(
            writer,
            sheet_name="Target Progress",
            index=False,
        )

        if not registration_mix_table.empty:
            registration_mix_table.to_excel(
                writer,
                sheet_name="Registration Mix",
            )

        if not corporate_category_display.empty:
            corporate_category_display.to_excel(
                writer,
                sheet_name="Corporate Breakdown",
            )

        if not addon_summary_table.empty:
            addon_summary_table.to_excel(
                writer,
                sheet_name="Add-On Purchases",
            )

        for addon_label, (
            breakdown_table,
            summary_table,
        ) in addon_market_category_trackers.items():
            short_name = ADDON_SHEET_SHORT_NAMES.get(
                addon_label,
                addon_label.replace(" ", "")[:15],
            )

            if not breakdown_table.empty:
                breakdown_table.to_excel(
                    writer,
                    # Excel sheet names cap at 31 characters.
                    sheet_name=(
                        f"AddOns {short_name}"[:31]
                    ),
                )

            if not summary_table.empty:
                summary_table.to_excel(
                    writer,
                    sheet_name=(
                        f"AddOns {short_name} Summary"[:31]
                    ),
                )

        if not country_market_table.empty:
            country_market_table.to_excel(
                writer,
                sheet_name="SG vs Non-SG (Country)",
            )

        if not nationality_market_table.empty:
            nationality_market_table.to_excel(
                writer,
                sheet_name="SG vs Non-SG (Nationality)",
            )

        if not age_gender_table.empty:
            age_gender_table.to_excel(
                writer,
                sheet_name="Age & Gender Splits",
            )

        if not category_gender_table.empty:
            category_gender_table.to_excel(
                writer,
                sheet_name="Category & Gender Splits",
            )

        if not age_category_gender_grid.empty:
            age_category_gender_grid.to_excel(
                writer,
                sheet_name="Age x Category x Gender",
                index=False,
            )

        if not capacity_segment_table.empty:
            capacity_segment_table.to_excel(
                writer,
                sheet_name="Capacity by Segment",
            )

        if not complimentary_category_display.empty:
            complimentary_category_display.to_excel(
                writer,
                sheet_name="Complimentary Breakdown",
            )

        for channel_label, channel_tracker in (
            channel_trackers.items()
        ):
            if channel_tracker.empty:
                continue

            short_channel = (
                channel_label.split(" ")[0]
            )

            channel_tracker.to_excel(
                writer,
                sheet_name=(
                    f"Tracker {short_channel}"[:31]
                ),
            )

        for channel_label, channel_tracker in (
            channel_trackers_daily.items()
        ):
            if channel_tracker.empty:
                continue

            short_channel = (
                channel_label.split(" ")[0]
            )

            channel_tracker.to_excel(
                writer,
                sheet_name=(
                    f"Tracker {short_channel} Daily"[:31]
                ),
            )

        daily_summary.to_excel(
            writer,
            sheet_name="Daily Registrations",
            index=False,
        )

        age_summary.to_excel(
            writer,
            sheet_name="Age Analysis",
            index=False,
        )

        gender_summary.to_excel(
            writer,
            sheet_name="Gender Analysis",
            index=False,
        )

        country_summary.to_excel(
            writer,
            sheet_name="Country Analysis",
            index=False,
        )

        quality_summary.to_excel(
            writer,
            sheet_name="Data Quality",
            index=False,
        )

        tracker_table.to_excel(
            writer,
            sheet_name="Registration Tracker",
            index=True,
            index_label="Category",
        )

        filtered_data.to_excel(
            writer,
            sheet_name="Filtered Data",
            index=False,
        )

        original_data.to_excel(
            writer,
            sheet_name="Original Source Data",
            index=False,
        )

        workbook = writer.book

        standard_worksheet_names = [
            "Category Summary",
            "Category Pace",
            "Target Progress",
            "Registration Mix",
            "Corporate Breakdown",
            "Add-On Purchases",
            "SG vs Non-SG (Country)",
            "SG vs Non-SG (Nationality)",
            "Age & Gender Splits",
            "Category & Gender Splits",
            "Age x Category x Gender",
            "Capacity by Segment",
            "Complimentary Breakdown",
            "Daily Registrations",
            "Age Analysis",
            "Gender Analysis",
            "Country Analysis",
            "Data Quality",
            "Filtered Data",
            "Original Source Data",
        ] + [
            sheet_name
            for addon_label in addon_market_category_trackers
            for sheet_name in (
                f"AddOns {ADDON_SHEET_SHORT_NAMES.get(addon_label, addon_label.replace(' ', '')[:15])}"[:31],
                f"AddOns {ADDON_SHEET_SHORT_NAMES.get(addon_label, addon_label.replace(' ', '')[:15])} Summary"[:31],
            )
        ] + [
            f"Tracker {channel_label.split(' ')[0]}"[:31]
            for channel_label in channel_trackers
        ] + [
            f"Tracker {channel_label.split(' ')[0]} Daily"[:31]
            for channel_label in channel_trackers_daily
        ]

        for worksheet_name in standard_worksheet_names:
            if worksheet_name in workbook.sheetnames:
                format_standard_worksheet(
                    workbook[worksheet_name]
                )

        format_tracker_worksheet(
            workbook["Registration Tracker"]
        )

        daily_sheet = workbook["Daily Registrations"]

        for cell in daily_sheet["A"][1:]:
            cell.number_format = "dd-mmm-yyyy"

        category_summary_sheet = workbook["Category Summary"]

        for cell in category_summary_sheet["C"][1:]:
            cell.number_format = "0.0%"

        category_pace_sheet = workbook["Category Pace"]

        for cell in category_pace_sheet["F"][1:]:
            cell.number_format = "0.0%"

    output.seek(0)

    return output


# =========================================================
# DISPLAY HELPERS
# =========================================================

def format_tracker_for_display(tracker):
    """
    Convert datetime column names into readable date labels.
    """
    display = tracker.copy()

    display.columns = [
        pd.Timestamp(column).strftime("%d-%b-%Y")
        for column in display.columns
    ]

    return display


def show_unavailable_message(field_name):
    """
    Display a consistent message when an optional field is unavailable.
    """
    st.info(
        f"{field_name} analysis is unavailable because the relevant "
        "column was not selected or was not found in the uploaded file."
    )


# =========================================================
# FILE UPLOAD
# =========================================================

uploaded_file = st.file_uploader(
    "Upload the registration CSV file",
    type=["csv"],
    help=(
        "The expected fields are Registration Date, Current Age, "
        "Gender, Country and Category Name."
    ),
)

if uploaded_file is None:
    st.info(
        "Upload a registration CSV file to begin."
    )
    st.stop()


# =========================================================
# READ FILE
# =========================================================

try:
    source_df = read_csv_file(uploaded_file)

except Exception as error:
    st.error(
        f"Unable to read the uploaded CSV file: {error}"
    )
    st.stop()

if source_df.empty:
    st.warning(
        "The uploaded CSV file does not contain any records."
    )
    st.stop()


# =========================================================
# COLUMN CONFIGURATION
# =========================================================

detected_date = detect_column(
    source_df.columns,
    COLUMN_ALIASES["registration_date"],
)

detected_age = detect_column(
    source_df.columns,
    COLUMN_ALIASES["age"],
)

detected_gender = detect_column(
    source_df.columns,
    COLUMN_ALIASES["gender"],
)

detected_country = detect_column(
    source_df.columns,
    COLUMN_ALIASES["country"],
)

detected_nationality = detect_column(
    source_df.columns,
    COLUMN_ALIASES["nationality"],
)

detected_category = detect_column(
    source_df.columns,
    COLUMN_ALIASES["category"],
)

detected_group_corporate = detect_column(
    source_df.columns,
    COLUMN_ALIASES["group_corporate"],
)

detected_addon_bag = detect_column(
    source_df.columns,
    COLUMN_ALIASES["addon_bag"],
)

detected_addon_itab = detect_column(
    source_df.columns,
    COLUMN_ALIASES["addon_itab"],
)

detected_addon_runflex = detect_column(
    source_df.columns,
    COLUMN_ALIASES["addon_runflex"],
)

with st.expander(
    "Source-column configuration",
    expanded=False,
):
    st.write(
        "The application attempts to identify the required columns "
        "automatically. Adjust the selections where necessary."
    )

    column_1, column_2 = st.columns(2)

    with column_1:
        registration_date_column = optional_column_selector(
            "Registration date column",
            source_df.columns,
            detected_date,
        )

        age_column = optional_column_selector(
            "Current age column",
            source_df.columns,
            detected_age,
        )

        gender_column = optional_column_selector(
            "Gender column",
            source_df.columns,
            detected_gender,
        )

    with column_2:
        country_column = optional_column_selector(
            "Country column",
            source_df.columns,
            detected_country,
            help_text=(
                "Country of residence/registration. Used for the "
                "Local vs International split."
            ),
        )

        nationality_column = optional_column_selector(
            "Nationality column",
            source_df.columns,
            detected_nationality,
            help_text=(
                "Optional. If the file has a separate "
                "nationality field, it enables a second, "
                "nationality-based Local vs International split "
                "on the Executive Snapshot."
            ),
        )

        category_column = optional_column_selector(
            "Category name column",
            source_df.columns,
            detected_category,
        )

        group_corporate_column = optional_column_selector(
            "Group / corporate name column",
            source_df.columns,
            detected_group_corporate,
            help_text=(
                "Optional. Values prefixed COMPLIMENTARY_ or "
                "GROUP_REGISTRATION_ enable the Corporate & "
                "Comps analysis."
            ),
        )

        addon_bag_column = optional_column_selector(
            "Gear Deposit Bag add-on column",
            source_df.columns,
            detected_addon_bag,
            help_text=(
                "Optional. A 0/1 purchase-flag column."
            ),
        )

        addon_itab_column = optional_column_selector(
            "iTab add-on column",
            source_df.columns,
            detected_addon_itab,
            help_text=(
                "Optional. A 0/1 purchase-flag column."
            ),
        )

        addon_runflex_column = optional_column_selector(
            "RunFlex add-on column",
            source_df.columns,
            detected_addon_runflex,
            help_text=(
                "Optional. A 0/1 purchase-flag column."
            ),
        )

        day_first = st.checkbox(
            "Dates use day-month-year format",
            value=True,
            help=(
                "Keep this selected for dates such as "
                "27/04/2026 or 27/4/2026 10:09."
            ),
        )

if registration_date_column is None:
    st.error(
        "A registration date column must be selected."
    )
    st.stop()

if category_column is None:
    st.error(
        "A category column must be selected."
    )
    st.stop()


# =========================================================
# PREPARE DATA
# =========================================================

try:
    addon_columns = {
        ADDON_LABELS["addon_bag"]: addon_bag_column,
        ADDON_LABELS["addon_itab"]: addon_itab_column,
        ADDON_LABELS["addon_runflex"]: addon_runflex_column,
    }

    addon_columns = {
        label: column
        for label, column in addon_columns.items()
        if column is not None
    }

    prepared_df = prepare_registration_data(
        dataframe=source_df,
        registration_date_column=registration_date_column,
        category_column=category_column,
        age_column=age_column,
        gender_column=gender_column,
        country_column=country_column,
        nationality_column=nationality_column,
        group_corporate_column=group_corporate_column,
        addon_columns=addon_columns,
        day_first=day_first,
    )

except Exception as error:
    st.error(
        f"Unable to prepare the uploaded data: {error}"
    )
    st.stop()


# =========================================================
# SIDEBAR FILTERS
# =========================================================

st.sidebar.header("Dashboard Filters")

valid_dates = prepared_df[
    "Registration Date Only"
].dropna()

if not valid_dates.empty:
    minimum_date = valid_dates.min().date()
    maximum_date = valid_dates.max().date()

    selected_date_range = st.sidebar.date_input(
        "Registration date range",
        value=(minimum_date, maximum_date),
        min_value=minimum_date,
        max_value=maximum_date,
    )
else:
    selected_date_range = None

available_categories = [
    category
    for category in CATEGORY_ORDER
    if category in prepared_df[
        "Grouped Category"
    ].unique()
]

selected_categories = st.sidebar.multiselect(
    "Race categories",
    options=available_categories,
    default=available_categories,
)

available_genders = sorted(
    prepared_df["Gender Clean"]
    .dropna()
    .unique()
    .tolist()
)

selected_genders = st.sidebar.multiselect(
    "Gender",
    options=available_genders,
    default=available_genders,
)

available_age_groups = [
    age_group
    for age_group in AGE_GROUP_ORDER
    if age_group in prepared_df["Age Group"].unique()
]

selected_age_groups = st.sidebar.multiselect(
    "Age groups",
    options=available_age_groups,
    default=available_age_groups,
)

available_markets = sorted(
    prepared_df["Market"]
    .dropna()
    .unique()
    .tolist()
)

selected_markets = st.sidebar.multiselect(
    "Market",
    options=available_markets,
    default=available_markets,
)

available_countries = sorted(
    prepared_df["Country Clean"]
    .dropna()
    .unique()
    .tolist()
)

selected_countries = st.sidebar.multiselect(
    "Countries",
    options=available_countries,
    default=available_countries,
)

filtered_df = apply_filters(
    data=prepared_df,
    selected_date_range=selected_date_range,
    selected_categories=selected_categories,
    selected_genders=selected_genders,
    selected_age_groups=selected_age_groups,
    selected_countries=selected_countries,
    selected_markets=selected_markets,
)


# =========================================================
# CREATE ANALYTICAL TABLES
# =========================================================

kpis = calculate_registration_kpis(filtered_df)

daily_summary = create_daily_registration_summary(
    filtered_df
)

category_summary = create_category_summary(
    filtered_df
)

category_pace = create_category_pace_table(
    filtered_df
)

age_summary = create_age_summary(
    filtered_df
)

gender_summary = create_gender_summary(
    filtered_df
)

country_summary = create_country_summary(
    filtered_df
)

quality_summary = create_data_quality_summary(
    prepared_df,
    age_available=age_column is not None,
    gender_available=gender_column is not None,
    country_available=country_column is not None,
)

tracker_table = create_tracker_table(
    filtered_df
)

target_progress = calculate_target_progress(
    filtered_df
)

registration_mix_table = create_registration_mix_table(
    filtered_df
)

complimentary_daily_table = (
    create_complimentary_daily_table(filtered_df)
)

corporate_category_counts, corporate_category_display = (
    create_corporate_category_table(filtered_df)
)

addon_labels_present = list(addon_columns.keys())

addon_summary_table = create_addon_summary_table(
    filtered_df, addon_labels_present
)

addon_market_category_trackers = (
    create_addon_market_category_trackers(
        filtered_df, addon_labels_present
    )
)

country_market_table = create_market_split_table(
    filtered_df, "Market", "Country"
)

nationality_market_table = (
    create_market_split_table(
        filtered_df, "Nationality Market", "Nationality"
    )
    if nationality_column
    else pd.DataFrame()
)

age_gender_table = create_age_gender_table(filtered_df)

category_gender_table = create_category_gender_table(filtered_df)

age_category_gender_grid = create_age_category_gender_grid(
    filtered_df
)

capacity_segment_table = create_capacity_segment_table(
    filtered_df,
    group_corporate_column=group_corporate_column,
    category_column=category_column,
)

channel_trackers = create_all_channel_trackers(
    filtered_df, mode="cumulative"
)

channel_trackers_daily = create_all_channel_trackers(
    filtered_df, mode="daily"
)

complimentary_category_counts, complimentary_category_display = (
    create_complimentary_category_table(filtered_df)
)

excel_report_bytes = create_excel_report(
    original_data=source_df,
    filtered_data=filtered_df,
    category_summary=category_summary,
    category_pace=category_pace,
    daily_summary=daily_summary,
    age_summary=age_summary,
    gender_summary=gender_summary,
    country_summary=country_summary,
    quality_summary=quality_summary,
    tracker_table=tracker_table,
    target_progress=target_progress,
    registration_mix_table=registration_mix_table,
    corporate_category_display=corporate_category_display,
    addon_summary_table=addon_summary_table,
    addon_market_category_trackers=addon_market_category_trackers,
    country_market_table=country_market_table,
    nationality_market_table=nationality_market_table,
    age_gender_table=age_gender_table,
    category_gender_table=category_gender_table,
    age_category_gender_grid=age_category_gender_grid,
    capacity_segment_table=capacity_segment_table,
    channel_trackers=channel_trackers,
    channel_trackers_daily=channel_trackers_daily,
    complimentary_category_display=complimentary_category_display,
)

archive_csv_bytes = create_archive_csv_bundle(
    tables={
        "SG vs Non-SG (Country)": country_market_table,
        "SG vs Non-SG (Nationality)": nationality_market_table,
        "Age & Gender Splits": age_gender_table,
        "Category & Gender Splits": category_gender_table,
        "Age x Category x Gender": age_category_gender_grid,
        "Capacity by Segment": capacity_segment_table,
        "Corporate Utilisation": corporate_category_counts,
        "Complimentary Utilisation": (
            complimentary_category_counts
        ),
        "Category Summary": category_summary,
        "Target Progress": target_progress,
        "Tracker - Retail (Cumulative)": channel_trackers.get(
            "Retail (excl. corporate & complimentary)"
        ),
        "Tracker - Corporate (Cumulative)": (
            channel_trackers.get("Corporate")
        ),
        "Tracker - Complimentary (Cumulative)": (
            channel_trackers.get("Complimentary")
        ),
        "Tracker - Retail (Daily)": (
            channel_trackers_daily.get(
                "Retail (excl. corporate & complimentary)"
            )
        ),
        "Tracker - Corporate (Daily)": (
            channel_trackers_daily.get("Corporate")
        ),
        "Tracker - Complimentary (Daily)": (
            channel_trackers_daily.get("Complimentary")
        ),
    }
)

original_filename = Path(uploaded_file.name).stem


# =========================================================
# DASHBOARD TABS
# =========================================================

(
    snapshot_tab,
    overview_tab,
    trends_tab,
    category_tab,
    demographics_tab,
    country_tab,
    corporate_tab,
    timing_tab,
    quality_tab,
    tracker_tab,
) = st.tabs(
    [
        "Executive Snapshot",
        "Executive Overview",
        "Registration Trends",
        "Category Performance",
        "Participant Demographics",
        "Country Analysis",
        "Corporate, Comps & Add-Ons",
        "Registration Timing",
        "Data Quality",
        "Registration Tracker",
    ]
)


# =========================================================
# TAB 0: EXECUTIVE SNAPSHOT
# =========================================================

with snapshot_tab:
    st.subheader("Executive Snapshot")

    latest_snapshot_date = (
        kpis["latest_date"].strftime("%d %b %Y")
        if kpis["latest_date"] is not None
        else "the latest upload"
    )

    st.caption(
        f"As of {latest_snapshot_date} — the same figures "
        "management reviews weekly, in one place."
    )

    snapshot_download_col_1, snapshot_download_col_2 = (
        st.columns(2)
    )

    with snapshot_download_col_1:
        st.download_button(
            label="Download Full Report (Excel)",
            data=excel_report_bytes,
            file_name=(
                f"{original_filename}_registration_analysis.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
            key="snapshot_download_excel",
        )

    with snapshot_download_col_2:
        st.download_button(
            label="Download All Tables (CSV)",
            data=archive_csv_bytes,
            file_name=(
                f"{original_filename}_archive_tables.csv"
            ),
            mime="text/csv",
            use_container_width=True,
            key="snapshot_download_csv",
        )

    st.caption(
        "Both buttons bundle every table across the Executive "
        "Snapshot, Executive Overview and Registration Tracker "
        "into one file — the Excel version keeps each table on "
        "its own sheet; the CSV version stacks them one after "
        "another with a header row naming each table, for a "
        "plain-text archive baseline."
    )

    if filtered_df.empty:
        st.warning(
            "No records match the selected filters."
        )
    else:
        st.markdown("#### SG vs Non-SG")

        market_col_1, market_col_2 = st.columns(2)

        with market_col_1:
            st.markdown("**By Country**")

            if country_market_table.empty:
                st.info("No country data available.")
            else:
                st.dataframe(
                    country_market_table,
                    use_container_width=True,
                )

        with market_col_2:
            st.markdown("**By Nationality**")

            if nationality_market_table.empty:
                st.info(
                    "No nationality column was selected for "
                    "this file, so this split is not available."
                )
            else:
                st.dataframe(
                    nationality_market_table,
                    use_container_width=True,
                )

        st.markdown("#### Age & Gender Splits")

        if age_gender_table.empty:
            st.info("No age/gender data available.")
        else:
            st.dataframe(
                age_gender_table,
                use_container_width=True,
            )

        st.markdown("#### Category & Gender Splits")

        if category_gender_table.empty:
            st.info("No category/gender data available.")
        else:
            st.dataframe(
                category_gender_table,
                use_container_width=True,
            )

        st.markdown("#### Age Range x Category x Gender")

        if age_category_gender_grid.empty:
            st.info("No data available for this breakdown.")
        else:
            st.dataframe(
                age_category_gender_grid,
                use_container_width=True,
                hide_index=True,
            )

            st.caption(
                "Each row is one category within one age band; "
                "percentages are of the overall campaign total. "
                "The '<Age Band> Total' row closes out each band."
            )

        st.markdown("#### Capacity by Segment")

        if capacity_segment_table.empty:
            st.info(
                "No mapped-category data available for the "
                "capacity breakdown."
            )
        else:
            st.dataframe(
                capacity_segment_table,
                use_container_width=True,
            )

            st.caption(
                "These are actual registered/sold counts by "
                "channel segment — Slots and Remaining are not "
                "shown here, since those are planning targets "
                "you maintain separately, not something the "
                "registration export can supply. Elite, National "
                "Championship and Score Collab are included as "
                "placeholder rows at 0 until there is a "
                "reliable rule (a dedicated field or promo code) "
                "to identify them in the export. The Corporate "
                "Registration row currently combines all "
                "non-staff corporate channels, including "
                "Standard Chartered's own corporate batches — "
                "let me know if that should split out further."
            )

        if group_corporate_column:
            st.markdown("#### Corporate Utilisation")

            if corporate_category_counts.empty:
                st.info(
                    "No group registrations in the current "
                    "selection."
                )
            else:
                st.dataframe(
                    corporate_category_counts,
                    use_container_width=True,
                )

                st.caption(
                    "Actual registrations per client per "
                    "category — entitlement/contracted numbers "
                    "are tracked in a separate sheet, so only "
                    "the utilisation side is shown here."
                )

            st.markdown("#### Complimentary Utilisation")

            if complimentary_category_counts.empty:
                st.info(
                    "No complimentary entries in the current "
                    "selection."
                )
            else:
                st.dataframe(
                    complimentary_category_counts,
                    use_container_width=True,
                )

                st.caption(
                    "Actual registrations per complimentary "
                    "programme per category, sorted by total "
                    "descending. Same shape as Corporate "
                    "Utilisation above, keyed on the "
                    "complimentary programme instead of the "
                    "corporate client."
                )


# =========================================================
# TAB 1: EXECUTIVE OVERVIEW
# =========================================================

with overview_tab:
    st.subheader("Executive Overview")

    st.caption(
        f"Registration closes "
        f"{REGISTRATION_CLOSE_DATE.strftime('%d %b %Y')} — "
        f"{kpis['days_to_close']:,} days remaining. "
        f"Race weekend: 4–6 Dec 2026."
    )

    if filtered_df.empty:
        st.warning(
            "No records match the selected filters."
        )

    metric_1, metric_2, metric_3, metric_4 = st.columns(4)

    metric_1.metric(
        "Total Registrations",
        f"{kpis['total']:,}",
    )

    latest_date_label = (
        kpis["latest_date"].strftime("%d %b %Y")
        if kpis["latest_date"] is not None
        else "Not available"
    )

    metric_2.metric(
        f"Latest Day — {latest_date_label}",
        f"{kpis['latest_day']:,}",
        delta=(
            kpis["latest_day"]
            - kpis["same_weekday_prior"]
        ),
        help=(
            "Difference against the same weekday one week earlier, "
            "which removes weekday seasonality from the comparison."
        ),
    )

    metric_3.metric(
        "Latest 7 Days",
        f"{kpis['latest_7_days']:,}",
        delta=(
            f"{kpis['week_change_percentage']:.1f}% "
            "vs preceding 7 days"
        ),
    )

    metric_4.metric(
        "7-Day Daily Average",
        f"{kpis['seven_day_average']:,.1f}",
    )

    metric_5, metric_6, metric_7, metric_8 = st.columns(4)

    metric_5.metric(
        "Countries Represented",
        f"{kpis['countries']:,}",
    )

    metric_6.metric(
        "Largest Category",
        kpis["top_category"],
        delta=f"{kpis['top_category_count']:,} registrations",
        delta_color="off",
    )

    singapore_count = int(
        filtered_df["Market"]
        .eq("Singapore")
        .sum()
    )

    international_count = int(
        filtered_df["Market"]
        .eq("International")
        .sum()
    )

    metric_7.metric(
        "Singapore Participants",
        f"{singapore_count:,}",
    )

    metric_8.metric(
        "International Participants",
        f"{international_count:,}",
    )

    st.markdown("#### Progress to Target")

    progress_with_targets = target_progress[
        target_progress["Target"].notna()
    ].copy()

    if progress_with_targets.empty:
        st.info(
            "No target groups match the selected filters."
        )
    else:
        # Bullet-style chart: the full target is drawn as a light
        # background bar, current registrations sit on top coloured
        # by pacing status, and the projected close-date total is
        # marked with a diamond.
        progress_figure = go.Figure()

        group_order = progress_with_targets[
            "Target Group"
        ].tolist()[::-1]

        progress_figure.add_trace(
            go.Bar(
                x=progress_with_targets["Target"],
                y=progress_with_targets["Target Group"],
                orientation="h",
                marker_color="#E8ECF0",
                name="Target",
                hovertemplate=(
                    "%{y}<br>Target: %{x:,}<extra></extra>"
                ),
            )
        )

        progress_figure.add_trace(
            go.Bar(
                x=progress_with_targets["Current"],
                y=progress_with_targets["Target Group"],
                orientation="h",
                marker_color=[
                    STATUS_COLORS.get(status, NEUTRAL_COLOR)
                    for status in progress_with_targets["Status"]
                ],
                name="Current",
                text=[
                    f"{current:,} ({fill:.0f}%)"
                    for current, fill in zip(
                        progress_with_targets["Current"],
                        progress_with_targets["Fill %"],
                    )
                ],
                textposition="auto",
                hovertemplate=(
                    "%{y}<br>Current: %{x:,}<extra></extra>"
                ),
            )
        )

        progress_figure.add_trace(
            go.Scatter(
                x=progress_with_targets["Projected at Close"],
                y=progress_with_targets["Target Group"],
                mode="markers",
                marker={
                    "symbol": "diamond",
                    "size": 10,
                    "color": "#3B3F46",
                },
                name="Projected at close",
                hovertemplate=(
                    "%{y}<br>Projected at close: "
                    "%{x:,}<extra></extra>"
                ),
            )
        )

        progress_figure.update_layout(
            barmode="overlay",
            yaxis={
                "categoryorder": "array",
                "categoryarray": group_order,
            },
            xaxis_title="Registrations",
            yaxis_title="",
            legend={
                "orientation": "h",
                "y": -0.2,
            },
            height=340,
        )

        st.plotly_chart(
            progress_figure,
            use_container_width=True,
        )

        st.caption(
            "Bar colour reflects pacing status against the "
            f"{REGISTRATION_CLOSE_DATE.strftime('%d %b %Y')} close "
            "date: green = on track or met, amber = projected within "
            "90% of target, red = off track. The diamond marks the "
            "projected total at close from a day-of-week adjusted "
            "forecast: each remaining day is estimated at that "
            "weekday's recent average (last 28 days, launch day "
            "and the latest partial day excluded), so a quiet "
            "Sunday no longer drags the whole projection down."
        )

    left_chart, right_chart = st.columns(2)

    with left_chart:
        st.markdown("#### Cumulative Registration Trend")

        if daily_summary.empty:
            st.info(
                "No valid registration dates are available."
            )
        else:
            cumulative_chart_data = daily_summary.copy()

            cumulative_chart_data[
                "Cumulative Registrations"
            ] = cumulative_chart_data[
                "Daily Registrations"
            ].cumsum()

            figure = px.line(
                cumulative_chart_data,
                x="Date",
                y="Cumulative Registrations",
                markers=True,
                color_discrete_sequence=[ACCENT_COLOR],
            )

            projection_data = create_weekday_adjusted_projection(
                daily_summary
            )

            if not projection_data.empty:
                figure.add_trace(
                    go.Scatter(
                        x=projection_data["Date"],
                        y=projection_data["Upper"],
                        mode="lines",
                        line={"width": 0},
                        showlegend=False,
                        hoverinfo="skip",
                    )
                )

                figure.add_trace(
                    go.Scatter(
                        x=projection_data["Date"],
                        y=projection_data["Lower"],
                        mode="lines",
                        line={"width": 0},
                        fill="tonexty",
                        fillcolor="rgba(31, 78, 121, 0.12)",
                        name="95% confidence band",
                        hoverinfo="skip",
                    )
                )

                figure.add_trace(
                    go.Scatter(
                        x=projection_data["Date"],
                        y=projection_data[
                            "Projected Cumulative"
                        ],
                        mode="lines",
                        line={
                            "dash": "dash",
                            "color": PROJECTION_COLOR,
                        },
                        name="Weekday-adjusted forecast",
                    )
                )

            overall_target = sum(
                value
                for value in CATEGORY_TARGETS.values()
                if value
            )

            figure.add_hline(
                y=overall_target,
                line_dash="dot",
                line_color=TARGET_COLOR,
                annotation_text=(
                    f"Overall target {overall_target:,}"
                ),
                annotation_position="top left",
                annotation_font_color=TARGET_COLOR,
            )

            add_milestone_markers(figure)

            figure.update_layout(
                xaxis_title="Registration date",
                yaxis_title="Cumulative registrations",
                hovermode="x unified",
                showlegend=False,
            )

            st.plotly_chart(
                figure,
                use_container_width=True,
            )

    with right_chart:
        st.markdown("#### Registrations by Category")

        figure = px.bar(
            category_summary,
            x="Registrations",
            y="Category",
            orientation="h",
            text_auto=True,
        )

        # Grey is the default; only the largest category carries the
        # accent colour, so the headline is visible at a glance.
        maximum_registrations = category_summary[
            "Registrations"
        ].max()

        figure.update_traces(
            marker_color=[
                ACCENT_COLOR
                if value == maximum_registrations
                else NEUTRAL_COLOR
                for value in category_summary["Registrations"]
            ]
        )

        figure.update_layout(
            xaxis_title="Registrations",
            yaxis_title="",
            yaxis={
                "categoryorder": "array",
                "categoryarray": CATEGORY_ORDER[::-1],
            },
        )

        st.plotly_chart(
            figure,
            use_container_width=True,
        )

    st.markdown("#### Management Observation")

    if kpis["latest_7_days"] > kpis["previous_7_days"]:
        direction_text = "increased"
    elif kpis["latest_7_days"] < kpis["previous_7_days"]:
        direction_text = "decreased"
    else:
        direction_text = "remained unchanged"

    observation_sentences = [
        f"Registration pace {direction_text} by "
        f"{abs(kpis['week_change_percentage']):.1f}% in the latest "
        f"seven-day period compared with the preceding seven days.",
        f"The largest category is {kpis['top_category']} with "
        f"{kpis['top_category_count']:,} registrations.",
    ]

    overview_pattern = summarise_daily_pattern(
        daily_summary
    )

    if (
        overview_pattern
        and overview_pattern["median"] is not None
        and overview_pattern["days_in_window"] > 0
    ):
        observation_sentences.append(
            f"{overview_pattern['above_median_last_14']} of the "
            f"last {overview_pattern['days_in_window']} days beat "
            "the campaign median of "
            f"{overview_pattern['median']:,.0f} registrations per "
            "day (launch day excluded) — a quick read on whether "
            "momentum is holding."
        )

    if (
        overview_pattern
        and overview_pattern["launch_count"]
        and kpis["total"] > 0
    ):
        launch_share = (
            overview_pattern["launch_count"]
            / kpis["total"]
            * 100
        )

        observation_sentences.append(
            "The "
            f"{REGISTRATION_OPEN_DATE.strftime('%d %b')} launch "
            f"day contributed {launch_share:.0f}% of all "
            "registrations to date and is treated as a known "
            "one-off, not a performance benchmark."
        )

    if not progress_with_targets.empty:
        behind_groups = progress_with_targets[
            progress_with_targets["Status"].isin(
                ["At risk", "Off track"]
            )
        ]["Target Group"].tolist()

        if behind_groups:
            observation_sentences.append(
                "Projected to miss target at the current run rate: "
                + ", ".join(behind_groups)
                + f" ({kpis['days_to_close']:,} days remain before "
                "registration closes)."
            )
        else:
            observation_sentences.append(
                "All target groups are pacing to meet their targets "
                "before registration closes on "
                f"{REGISTRATION_CLOSE_DATE.strftime('%d %b %Y')}."
            )

    st.info(" ".join(observation_sentences))


# =========================================================
# TAB 2: REGISTRATION TRENDS
# =========================================================

with trends_tab:
    st.subheader("Registration Trends")

    if daily_summary.empty:
        st.info(
            "No valid registration dates are available for trend analysis."
        )
    else:
        daily_pattern = summarise_daily_pattern(
            daily_summary
        )

        daily_figure = px.bar(
            daily_summary,
            x="Date",
            y="Daily Registrations",
            text_auto=False,
            title="Daily Registrations versus Campaign Median",
        )

        # Colour logic: the launch-day surge is a known structural
        # outlier (purple, labelled); days above the campaign
        # median carry the accent colour; ordinary days stay grey.
        # The median excludes launch day so it reflects genuine
        # day-to-day demand.
        median_value = (
            daily_pattern["median"]
            if daily_pattern and daily_pattern["median"]
            else None
        )

        bar_colors = []

        for _, day_row in daily_summary.iterrows():
            if day_row["Date"] == REGISTRATION_OPEN_DATE:
                bar_colors.append(LAUNCH_COLOR)
            elif (
                median_value is not None
                and day_row["Daily Registrations"]
                > median_value
            ):
                bar_colors.append(ACCENT_COLOR)
            else:
                bar_colors.append(NEUTRAL_COLOR)

        daily_figure.update_traces(
            marker_color=bar_colors,
        )

        if median_value is not None:
            daily_figure.add_hline(
                y=median_value,
                line_dash="dot",
                line_color="#3B3F46",
                annotation_text=(
                    f"Campaign median {median_value:,.0f}/day "
                    "(excl. launch)"
                ),
                annotation_position="top left",
            )

        if daily_pattern and daily_pattern["launch_count"]:
            daily_figure.add_annotation(
                x=REGISTRATION_OPEN_DATE,
                y=daily_pattern["launch_count"],
                text=(
                    "Launch day: "
                    f"{daily_pattern['launch_count']:,}"
                ),
                showarrow=True,
                arrowhead=2,
                ay=-30,
                font={"color": LAUNCH_COLOR},
                arrowcolor=LAUNCH_COLOR,
            )

        if daily_pattern and daily_pattern["best_date"] is not None:
            daily_figure.add_annotation(
                x=daily_pattern["best_date"],
                y=daily_pattern["best_count"],
                text=(
                    "Best campaign day: "
                    f"{daily_pattern['best_count']:,} "
                    f"({daily_pattern['best_date'].strftime('%d %b')})"
                ),
                showarrow=True,
                arrowhead=2,
                ay=-45,
                font={"color": ACCENT_COLOR},
                arrowcolor=ACCENT_COLOR,
            )

        add_milestone_markers(daily_figure)

        daily_figure.update_layout(
            xaxis_title="Registration date",
            yaxis_title="Daily registrations",
            hovermode="x unified",
        )

        st.plotly_chart(
            daily_figure,
            use_container_width=True,
        )

        st.caption(
            "How to read this: blue days beat the typical "
            "campaign day; grey days did not. The launch-day "
            f"surge on {REGISTRATION_OPEN_DATE.strftime('%d %b')} "
            "(purple) is expected and is excluded from the median "
            "and the best-day call-out. Blue days that do not "
            "coincide with a marked milestone are worth "
            "investigating — something drove demand that day."
        )

        cumulative_data = daily_summary.copy()

        cumulative_data[
            "Cumulative Registrations"
        ] = cumulative_data[
            "Daily Registrations"
        ].cumsum()

        cumulative_figure = px.line(
            cumulative_data,
            x="Date",
            y="Cumulative Registrations",
            markers=True,
            title=(
                "Cumulative Registrations and Run-Rate Projection "
                f"to {REGISTRATION_CLOSE_DATE.strftime('%d %b %Y')}"
            ),
            color_discrete_sequence=[ACCENT_COLOR],
        )

        trend_projection = create_weekday_adjusted_projection(
            daily_summary
        )

        if not trend_projection.empty:
            cumulative_figure.add_trace(
                go.Scatter(
                    x=trend_projection["Date"],
                    y=trend_projection["Upper"],
                    mode="lines",
                    line={"width": 0},
                    showlegend=False,
                    hoverinfo="skip",
                )
            )

            cumulative_figure.add_trace(
                go.Scatter(
                    x=trend_projection["Date"],
                    y=trend_projection["Lower"],
                    mode="lines",
                    line={"width": 0},
                    fill="tonexty",
                    fillcolor="rgba(31, 78, 121, 0.12)",
                    name="95% confidence band",
                    hoverinfo="skip",
                )
            )

            cumulative_figure.add_trace(
                go.Scatter(
                    x=trend_projection["Date"],
                    y=trend_projection[
                        "Projected Cumulative"
                    ],
                    mode="lines",
                    line={
                        "dash": "dash",
                        "color": PROJECTION_COLOR,
                    },
                    name="Weekday-adjusted forecast",
                )
            )

        overall_target = sum(
            value
            for value in CATEGORY_TARGETS.values()
            if value
        )

        cumulative_figure.add_hline(
            y=overall_target,
            line_dash="dot",
            line_color=TARGET_COLOR,
            annotation_text=(
                f"Overall target {overall_target:,}"
            ),
            annotation_position="top left",
            annotation_font_color=TARGET_COLOR,
        )

        add_milestone_markers(cumulative_figure)

        cumulative_figure.update_layout(
            xaxis_title="Registration date",
            yaxis_title="Cumulative registrations",
            hovermode="x unified",
        )

        st.plotly_chart(
            cumulative_figure,
            use_container_width=True,
        )

        st.caption(
            "The dashed line is a day-of-week adjusted forecast: "
            "each remaining day is estimated at that weekday's "
            "average over the last 28 days (launch day excluded). "
            "The shaded band is the 95% confidence range — the "
            "true outcome should land inside it 19 times out of "
            "20 if demand patterns hold. A target line sitting "
            "above the band's upper edge means the target is "
            "unlikely without intervention, not just behind."
        )

        rolling_data = daily_summary.copy()

        rolling_data[
            "7-Day Moving Average"
        ] = rolling_data[
            "Daily Registrations"
        ].rolling(
            window=7,
            min_periods=1,
        ).mean()

        # Same visual grammar as the other indicator charts: a bar
        # that beats its own 7-day moving average carries the
        # accent colour. Runs of blue mean demand is accelerating
        # past its recent trend; runs of grey mean it is fading
        # below trend. Launch day keeps its outlier colour.
        rolling_colors = []

        for _, rolling_row in rolling_data.iterrows():
            if rolling_row["Date"] == REGISTRATION_OPEN_DATE:
                rolling_colors.append(LAUNCH_COLOR)
            elif (
                rolling_row["Daily Registrations"]
                > rolling_row["7-Day Moving Average"]
            ):
                rolling_colors.append(ACCENT_COLOR)
            else:
                rolling_colors.append(NEUTRAL_COLOR)

        rolling_figure = go.Figure()

        rolling_figure.add_trace(
            go.Bar(
                x=rolling_data["Date"],
                y=rolling_data["Daily Registrations"],
                name="Daily Registrations",
                marker_color=rolling_colors,
            )
        )

        rolling_figure.add_trace(
            go.Scatter(
                x=rolling_data["Date"],
                y=rolling_data["7-Day Moving Average"],
                mode="lines",
                name="7-Day Moving Average",
                line={
                    "color": "#3B3F46",
                    "width": 3,
                },
            )
        )

        add_milestone_markers(rolling_figure)

        rolling_figure.update_layout(
            title=(
                "Daily Registrations versus 7-Day "
                "Moving Average"
            ),
            xaxis_title="Registration date",
            yaxis_title="Registrations",
            hovermode="x unified",
        )

        st.plotly_chart(
            rolling_figure,
            use_container_width=True,
        )

        st.caption(
            "Blue bars beat their own 7-day moving average — a "
            "run of blue means demand is accelerating past its "
            "recent trend, while a run of grey means it is "
            "fading below trend. Watch for the switch: several "
            "grey days in a row after a blue stretch is the "
            "earliest visual sign that momentum is turning, "
            "before it shows up in the weekly totals."
        )

        trend_metric_1, trend_metric_2, trend_metric_3 = (
            st.columns(3)
        )

        if daily_pattern and daily_pattern["best_date"] is not None:
            trend_metric_1.metric(
                "Best Campaign Day (excl. launch)",
                daily_pattern["best_date"].strftime(
                    "%d %b %Y"
                ),
                delta=(
                    f"{daily_pattern['best_count']:,} "
                    "registrations"
                ),
                delta_color="off",
            )
        else:
            trend_metric_1.metric(
                "Best Campaign Day (excl. launch)",
                "Not available",
            )

        if daily_pattern and daily_pattern["launch_count"]:
            trend_metric_2.metric(
                "Launch Day "
                f"({REGISTRATION_OPEN_DATE.strftime('%d %b')})",
                f"{daily_pattern['launch_count']:,}",
                help=(
                    "The opening-day surge is a structural "
                    "one-off and is excluded from the campaign "
                    "median and best-day figures."
                ),
            )
        else:
            trend_metric_2.metric(
                "Launch Day "
                f"({REGISTRATION_OPEN_DATE.strftime('%d %b')})",
                "Not in filter",
            )

        if daily_pattern and daily_pattern["median"] is not None:
            trend_metric_3.metric(
                "Median per Day (excl. launch)",
                f"{daily_pattern['median']:,.0f}",
            )
        else:
            trend_metric_3.metric(
                "Median per Day (excl. launch)",
                "Not available",
            )


# =========================================================
# TAB 3: CATEGORY PERFORMANCE
# =========================================================

with category_tab:
    st.subheader("Category Performance")

    category_display = category_summary.copy()

    category_display["Share"] = (
        category_display["Share"]
        .map(lambda value: f"{value:.1f}%")
    )

    st.dataframe(
        category_display,
        use_container_width=True,
        hide_index=True,
    )

    category_figure = px.bar(
        category_summary,
        x="Registrations",
        y="Category",
        orientation="h",
        text="Registrations",
        title="Total Registrations by Summarised Category",
        color="Category",
        color_discrete_map=CATEGORY_COLORS,
    )

    category_figure.update_layout(
        xaxis_title="Registrations",
        yaxis_title="",
        yaxis={
            "categoryorder": "array",
            "categoryarray": CATEGORY_ORDER[::-1],
        },
        showlegend=False,
    )

    st.plotly_chart(
        category_figure,
        use_container_width=True,
    )

    st.markdown("#### Target Status")

    target_display = target_progress.copy()

    target_display["Target"] = target_display["Target"].map(
        lambda value: f"{int(value):,}"
        if pd.notna(value)
        else "—"
    )

    target_display["Fill %"] = target_display["Fill %"].map(
        lambda value: f"{value:.1f}%"
        if pd.notna(value)
        else "—"
    )

    target_display["Current"] = target_display[
        "Current"
    ].map(lambda value: f"{value:,}")

    target_display["Projected at Close"] = target_display[
        "Projected at Close"
    ].map(lambda value: f"{value:,}")

    st.dataframe(
        target_display,
        use_container_width=True,
        hide_index=True,
    )

    st.caption(
        "Projection = current registrations plus a day-of-week "
        "adjusted forecast of the remaining days to the "
        f"{REGISTRATION_CLOSE_DATE.strftime('%d %b %Y')} close: "
        "each remaining day is estimated at that weekday's recent "
        "average for the group (last 28 days, launch day "
        "excluded). The two Kids Dash 1.6KM categories share one "
        "combined target."
    )

    st.markdown("#### Latest Seven-Day Category Pace")

    pace_display = category_pace.copy()

    pace_display["Change %"] = (
        pace_display["Change %"]
        .map(lambda value: f"{value:.1f}%")
    )

    st.dataframe(
        pace_display,
        use_container_width=True,
        hide_index=True,
    )

    # Sort by the size of the change (not the fixed category order),
    # colour increases green and decreases red, and grey out any
    # category whose previous-week base is too small for its change
    # to be meaningful.
    pace_sorted = category_pace.sort_values(
        "Change"
    ).copy()

    pace_colors = []

    for _, pace_row in pace_sorted.iterrows():
        if pace_row["Previous 7 Days"] < SMALL_BASE_THRESHOLD:
            pace_colors.append(NEUTRAL_COLOR)
        elif pace_row["Change"] >= 0:
            pace_colors.append(POSITIVE_COLOR)
        else:
            pace_colors.append(NEGATIVE_COLOR)

    pace_chart = px.bar(
        pace_sorted,
        x="Change",
        y="Category",
        orientation="h",
        text=[
            f"{change:+,}"
            for change in pace_sorted["Change"]
        ],
        title=(
            "Change in Registrations: Latest 7 Days "
            "versus Preceding 7 Days"
        ),
    )

    pace_chart.update_traces(
        marker_color=pace_colors,
        textposition="outside",
    )

    pace_chart.add_vline(
        x=0,
        line_color="#3B3F46",
        line_width=1,
    )

    pace_chart.update_layout(
        xaxis_title="Change in registrations",
        yaxis_title="",
        yaxis={
            "categoryorder": "array",
            "categoryarray": pace_sorted["Category"].tolist(),
        },
    )

    st.caption(
        "Signed labels are kept on every bar so direction does not "
        "rely on colour alone. Categories with fewer than "
        f"{SMALL_BASE_THRESHOLD} registrations in the preceding week "
        "are shown in grey because percentage swings on small bases "
        "are not meaningful."
    )

    st.plotly_chart(
        pace_chart,
        use_container_width=True,
    )

    valid_category_trend = filtered_df[
        filtered_df["Registration Date Only"].notna()
        & filtered_df["Grouped Category"].ne("Unmapped")
    ].copy()

    if not valid_category_trend.empty:
        category_daily = (
            valid_category_trend.groupby(
                [
                    "Registration Date Only",
                    "Grouped Category",
                ],
                observed=True,
            )
            .size()
            .reset_index(name="Registrations")
        )

        category_daily[
            "Cumulative Registrations"
        ] = (
            category_daily
            .sort_values(
                [
                    "Grouped Category",
                    "Registration Date Only",
                ]
            )
            .groupby(
                "Grouped Category",
                observed=True,
            )["Registrations"]
            .cumsum()
        )

        category_trend_figure = px.line(
            category_daily,
            x="Registration Date Only",
            y="Cumulative Registrations",
            color="Grouped Category",
            color_discrete_map=CATEGORY_COLORS,
            title="Cumulative Registration Trend by Category",
        )

        add_milestone_markers(category_trend_figure)

        category_trend_figure.update_layout(
            xaxis_title="Registration date",
            yaxis_title="Cumulative registrations",
            legend_title="Category",
            hovermode="x unified",
        )

        st.plotly_chart(
            category_trend_figure,
            use_container_width=True,
        )


# =========================================================
# TAB 4: DEMOGRAPHICS
# =========================================================

with demographics_tab:
    st.subheader("Participant Demographics")

    demographic_column_1, demographic_column_2 = (
        st.columns(2)
    )

    with demographic_column_1:
        st.markdown("#### Age Profile")

        if age_column is None:
            show_unavailable_message("Age")
        else:
            age_figure = px.bar(
                age_summary,
                x="Age Group",
                y="Participants",
                text="Participants",
                color="Age Group",
                color_discrete_map=AGE_GROUP_COLORS,
            )

            age_figure.update_layout(
                xaxis_title="Age group",
                yaxis_title="Participants",
                showlegend=False,
            )

            st.plotly_chart(
                age_figure,
                use_container_width=True,
            )

            valid_ages = filtered_df[
                "Current Age Clean"
            ].dropna()

            valid_ages = valid_ages[
                valid_ages.between(
                    0,
                    110,
                    inclusive="both",
                )
            ]

            if not valid_ages.empty:
                age_metric_1, age_metric_2 = st.columns(2)

                age_metric_1.metric(
                    "Average Age",
                    f"{valid_ages.mean():.1f}",
                )

                age_metric_2.metric(
                    "Median Age",
                    f"{valid_ages.median():.1f}",
                )

    with demographic_column_2:
        st.markdown("#### Gender Profile")

        if gender_column is None:
            show_unavailable_message("Gender")
        else:
            gender_figure = px.pie(
                gender_summary,
                names="Gender",
                values="Participants",
                hole=0.45,
                color="Gender",
                color_discrete_map=GENDER_COLORS,
            )

            gender_figure.update_traces(
                textinfo="label+percent",
            )

            st.plotly_chart(
                gender_figure,
                use_container_width=True,
            )

    if age_column is not None or gender_column is not None:
        # Absolute stacked bars are dominated by the largest
        # categories; the share view makes each category's internal
        # mix directly comparable.
        breakdown_mode = st.radio(
            "Category breakdown display",
            options=[
                "Absolute counts",
                "Share of category (100%)",
            ],
            horizontal=True,
        )

        breakdown_norm = (
            "percent"
            if breakdown_mode == "Share of category (100%)"
            else None
        )

        breakdown_axis_title = (
            "Share of category (%)"
            if breakdown_norm
            else "Participants"
        )

    if age_column is not None:
        st.markdown("#### Age Group by Race Category")

        age_category = (
            filtered_df.groupby(
                [
                    "Grouped Category",
                    "Age Group",
                ],
                observed=True,
            )
            .size()
            .reset_index(name="Participants")
        )

        age_category = age_category[
            age_category[
                "Grouped Category"
            ].ne("Unmapped")
        ]

        age_category_figure = px.bar(
            age_category,
            x="Grouped Category",
            y="Participants",
            color="Age Group",
            color_discrete_map=AGE_GROUP_COLORS,
            category_orders={
                "Age Group": AGE_GROUP_ORDER,
                "Grouped Category": CATEGORY_ORDER,
            },
            barmode="stack",
        )

        if breakdown_norm:
            age_category_figure.update_layout(
                barnorm=breakdown_norm,
            )

        age_category_figure.update_layout(
            xaxis_title="Race category",
            yaxis_title=breakdown_axis_title,
            legend_title="Age group",
        )

        st.plotly_chart(
            age_category_figure,
            use_container_width=True,
        )

    if gender_column is not None:
        st.markdown("#### Gender by Race Category")

        gender_category = (
            filtered_df.groupby(
                [
                    "Grouped Category",
                    "Gender Clean",
                ],
                observed=True,
            )
            .size()
            .reset_index(name="Participants")
        )

        gender_category = gender_category[
            gender_category[
                "Grouped Category"
            ].ne("Unmapped")
        ]

        gender_category_figure = px.bar(
            gender_category,
            x="Grouped Category",
            y="Participants",
            color="Gender Clean",
            color_discrete_map=GENDER_COLORS,
            category_orders={
                "Grouped Category": CATEGORY_ORDER,
            },
            barmode="stack",
        )

        if breakdown_norm:
            gender_category_figure.update_layout(
                barnorm=breakdown_norm,
            )

        gender_category_figure.update_layout(
            xaxis_title="Race category",
            yaxis_title=breakdown_axis_title,
            legend_title="Gender",
        )

        st.plotly_chart(
            gender_category_figure,
            use_container_width=True,
        )


# =========================================================
# TAB 5: COUNTRY ANALYSIS
# =========================================================

with country_tab:
    st.subheader("Country and Market Analysis")

    if country_column is None:
        show_unavailable_message("Country")
    else:
        country_metric_1, country_metric_2, country_metric_3 = (
            st.columns(3)
        )

        known_country_data = filtered_df[
            filtered_df["Country Clean"].ne("Unknown")
        ]

        singapore_total = int(
            known_country_data["Market"]
            .eq("Singapore")
            .sum()
        )

        international_total = int(
            known_country_data["Market"]
            .eq("International")
            .sum()
        )

        known_total = len(known_country_data)

        international_share = (
            international_total / known_total * 100
            if known_total > 0
            else 0
        )

        country_metric_1.metric(
            "Countries Represented",
            f"{known_country_data['Country Clean'].nunique():,}",
        )

        country_metric_2.metric(
            "International Participants",
            f"{international_total:,}",
        )

        country_metric_3.metric(
            "International Share",
            f"{international_share:.1f}%",
        )

        top_country_count = st.slider(
            "Number of countries to display",
            min_value=5,
            max_value=min(
                30,
                max(
                    5,
                    len(country_summary),
                ),
            ),
            value=min(
                10,
                max(
                    5,
                    len(country_summary),
                ),
            ),
        )

        top_countries = (
            country_summary[
                country_summary["Country"].ne(
                    "Unknown"
                )
            ]
            .head(top_country_count)
        )

        # Pareto view: bars for the top countries plus a cumulative
        # share line, so statements like "the top five countries
        # cover 96% of participants" are visible at a glance.
        known_summary = country_summary[
            country_summary["Country"].ne("Unknown")
        ].copy()

        known_participant_total = known_summary[
            "Participants"
        ].sum()

        pareto_data = top_countries.copy()

        pareto_data["Cumulative Share"] = (
            pareto_data["Participants"].cumsum()
            / known_participant_total
            * 100
            if known_participant_total > 0
            else 0
        )

        other_count = int(
            known_participant_total
            - pareto_data["Participants"].sum()
        )

        country_figure = go.Figure()

        country_figure.add_trace(
            go.Bar(
                x=pareto_data["Country"],
                y=pareto_data["Participants"],
                text=pareto_data["Participants"],
                textposition="outside",
                marker_color=[
                    MARKET_COLORS["Singapore"]
                    if country == "Singapore"
                    else MARKET_COLORS["International"]
                    for country in pareto_data["Country"]
                ],
                name="Participants",
            )
        )

        country_figure.add_trace(
            go.Scatter(
                x=pareto_data["Country"],
                y=pareto_data["Cumulative Share"],
                mode="lines+markers",
                line={"color": "#3B3F46"},
                name="Cumulative share",
                yaxis="y2",
                hovertemplate=(
                    "%{x}<br>Cumulative share: "
                    "%{y:.1f}%<extra></extra>"
                ),
            )
        )

        country_figure.update_layout(
            title=(
                f"Top {top_country_count} Countries "
                "(Pareto view)"
            ),
            xaxis_title="",
            yaxis={"title": "Participants"},
            yaxis2={
                "title": "Cumulative share (%)",
                "overlaying": "y",
                "side": "right",
                "range": [0, 105],
                "showgrid": False,
            },
            legend={
                "orientation": "h",
                "y": -0.25,
            },
        )

        st.plotly_chart(
            country_figure,
            use_container_width=True,
        )

        if other_count > 0:
            remaining_countries = (
                len(known_summary) - len(pareto_data)
            )

            st.caption(
                f"A further {other_count:,} participants come from "
                f"the remaining {remaining_countries:,} countries."
            )

        market_summary = (
            filtered_df["Market"]
            .value_counts()
            .rename_axis("Market")
            .reset_index(name="Participants")
        )

        market_figure = px.pie(
            market_summary,
            names="Market",
            values="Participants",
            hole=0.45,
            title="Singapore versus International Participants",
            color="Market",
            color_discrete_map=MARKET_COLORS,
        )

        market_figure.update_traces(
            textinfo="label+percent",
        )

        st.plotly_chart(
            market_figure,
            use_container_width=True,
        )

        st.markdown("#### International Participation by Category")

        market_category = (
            filtered_df.groupby(
                [
                    "Grouped Category",
                    "Market",
                ],
                observed=True,
            )
            .size()
            .reset_index(name="Participants")
        )

        market_category = market_category[
            market_category[
                "Grouped Category"
            ].ne("Unmapped")
        ]

        market_category_figure = px.bar(
            market_category,
            x="Grouped Category",
            y="Participants",
            color="Market",
            color_discrete_map=MARKET_COLORS,
            category_orders={
                "Grouped Category": CATEGORY_ORDER,
            },
            barmode="stack",
        )

        market_category_figure.update_layout(
            xaxis_title="Race category",
            yaxis_title="Participants",
            legend_title="Market",
        )

        st.plotly_chart(
            market_category_figure,
            use_container_width=True,
        )


# =========================================================
# TAB 6: CORPORATE, COMPS AND ADD-ONS
# =========================================================

with corporate_tab:
    st.subheader("Corporate, Comps & Add-Ons")

    if group_corporate_column is None:
        st.info(
            "No Group/Corporate Name column was found in this "
            "file, so channel analysis is not available. Select "
            "the column manually under Source-column "
            "configuration if it exists under another name."
        )
    elif filtered_df.empty:
        st.warning(
            "No records match the selected filters."
        )
    else:
        channel_counts = filtered_df[
            filtered_df["Registration Date Only"].notna()
        ]["Registration Type"].value_counts()

        total_valid = int(channel_counts.sum())

        group_total = int(
            channel_counts.get("Group Registration", 0)
        )

        complimentary_total = int(
            channel_counts.get("Complimentary", 0)
        )

        other_total = int(
            channel_counts.get("Other", 0)
        )

        company_count = (
            filtered_df["Corporate Group"]
            .dropna()
            .nunique()
        )

        channel_metric_1, channel_metric_2, channel_metric_3 = (
            st.columns(3)
        )

        channel_metric_1.metric(
            "Group Registrations",
            f"{group_total:,}",
            delta=(
                f"{group_total / total_valid * 100:.1f}% "
                "of all registrations"
                if total_valid
                else None
            ),
            delta_color="off",
        )

        channel_metric_2.metric(
            "Complimentary Entries",
            f"{complimentary_total:,}",
            delta=(
                f"{complimentary_total / total_valid * 100:.1f}% "
                "of all registrations"
                if total_valid
                else None
            ),
            delta_color="off",
        )

        channel_metric_3.metric(
            "Corporate Groups",
            f"{company_count:,}",
        )

        if other_total:
            st.warning(
                f"{other_total:,} rows have a Group/Corporate "
                "value with neither the COMPLIMENTARY_ nor the "
                "GROUP_REGISTRATION_ prefix. They are shown as "
                "'Other' and are worth checking at source."
            )

        st.markdown("#### Daily Registration Mix by Channel")

        mix_chart_data = filtered_df[
            filtered_df["Registration Date Only"].notna()
        ]

        mix_daily = (
            mix_chart_data.groupby(
                [
                    "Registration Date Only",
                    "Registration Type",
                ],
                observed=True,
            )
            .size()
            .reset_index(name="Registrations")
        )

        mix_figure = px.bar(
            mix_daily,
            x="Registration Date Only",
            y="Registrations",
            color="Registration Type",
            color_discrete_map=REGISTRATION_TYPE_COLORS,
            category_orders={
                "Registration Type": (
                    REGISTRATION_TYPE_ORDER + ["Other"]
                ),
            },
            barmode="stack",
        )

        add_milestone_markers(mix_figure)

        mix_figure.update_layout(
            xaxis_title="Registration date",
            yaxis_title="Registrations",
            legend_title="Channel",
            hovermode="x unified",
        )

        st.plotly_chart(
            mix_figure,
            use_container_width=True,
        )

        st.caption(
            "Blue segments are corporate group registrations and "
            "purple segments are complimentary allocations. "
            "Corporate blocks landing on otherwise quiet days "
            "flatter the daily trend — this chart shows how much "
            "of any spike was organic public demand versus a "
            "single bulk upload."
        )

        if not registration_mix_table.empty:
            st.markdown("#### Daily Mix Table")

            st.dataframe(
                registration_mix_table,
                use_container_width=True,
            )

            st.caption(
                "Each column is a registration date; scroll "
                "horizontally for the full campaign. Daily Total "
                "and Cumulative Total are the bottom two rows."
            )

        st.markdown("#### Complimentary Entries by Programme")

        if complimentary_daily_table.empty:
            st.info(
                "No complimentary entries in the current "
                "selection."
            )
        else:
            programme_totals = (
                complimentary_daily_table
                .loc["Programme Total"]
                .drop("Daily Total")
                .sort_values()
            )

            programme_figure = px.bar(
                programme_totals.reset_index(),
                x="Programme Total",
                y="Complimentary Programme",
                orientation="h",
                text="Programme Total",
                title=(
                    "Total Complimentary Entries by Programme"
                ),
            )

            maximum_programme = programme_totals.max()

            programme_figure.update_traces(
                marker_color=[
                    "#7B2D8B"
                    if value == maximum_programme
                    else NEUTRAL_COLOR
                    for value in programme_totals
                ]
            )

            programme_figure.update_layout(
                xaxis_title="Complimentary entries",
                yaxis_title="",
            )

            st.plotly_chart(
                programme_figure,
                use_container_width=True,
            )

            st.dataframe(
                complimentary_daily_table,
                use_container_width=True,
            )

            st.caption(
                "Each column is a complimentary programme (the "
                "identity after the COMPLIMENTARY_ prefix). Use "
                "this to track burn against each sponsor's or "
                "partner's agreed allocation — a programme "
                "consuming entries faster than planned is a "
                "conversation to have early, not at close."
            )

        st.markdown("#### Corporate Category Mix")

        if corporate_category_display.empty:
            st.info(
                "No group registrations in the current "
                "selection."
            )
        else:
            corporate_share = (
                corporate_category_counts.drop(
                    columns="Total"
                )
            )

            top_companies = (
                corporate_category_counts.head(12).index
            )

            share_data = (
                corporate_share.loc[top_companies]
                .reset_index()
                .melt(
                    id_vars="Corporate Group",
                    var_name="Category",
                    value_name="Registrations",
                )
            )

            share_data = share_data[
                share_data["Registrations"] > 0
            ]

            share_figure = px.bar(
                share_data,
                x="Registrations",
                y="Corporate Group",
                color="Category",
                color_discrete_map=CATEGORY_COLORS,
                category_orders={
                    "Corporate Group": (
                        top_companies.tolist()[::-1]
                    ),
                    "Category": CATEGORY_ORDER,
                },
                orientation="h",
                barmode="stack",
                title=(
                    "Category Share Within Each Corporate "
                    "Group"
                ),
            )

            share_figure.update_layout(
                barnorm="percent",
                xaxis_title="Share of the company's registrations (%)",
                yaxis_title="",
                legend_title="Category",
                height=max(
                    340,
                    36 * len(top_companies) + 120,
                ),
            )

            st.plotly_chart(
                share_figure,
                use_container_width=True,
            )

            st.dataframe(
                corporate_category_display,
                use_container_width=True,
            )

            st.caption(
                "Each cell reads count (share of that company's "
                "registrations). Companies skewing to 10km and "
                "5km are wellness accounts — pitch team bundles "
                "and next year's early commitment. Companies "
                "with meaningful 42.195KM and 21.0975KM shares "
                "are performance accounts — pitch premium "
                "packages such as pace groups or hospitality. "
                "Batch entries (e.g. SCB Batch 1, Batch 2) are "
                "rolled up to one company."
            )

        st.markdown("#### Add-On Purchases")

        if not addon_labels_present:
            st.info(
                "No add-on purchase columns were found in this "
                "file. Select them manually under "
                "Source-column configuration if they exist "
                "under different names."
            )
        elif addon_summary_table.empty:
            st.info(
                "No add-on purchase data in the current "
                "selection."
            )
        else:
            addon_totals = (
                addon_summary_table["Total"]
                .drop("Total Add-Ons Purchased")
                .sort_values(ascending=False)
            )

            addon_metric_columns = st.columns(
                len(addon_totals)
            )

            total_registrations_in_view = int(
                filtered_df[
                    "Registration Date Only"
                ].notna().sum()
            )

            for metric_column, (
                addon_label,
                addon_total,
            ) in zip(
                addon_metric_columns,
                addon_totals.items(),
            ):
                take_up_rate = (
                    addon_total
                    / total_registrations_in_view
                    * 100
                    if total_registrations_in_view
                    else 0
                )

                metric_column.metric(
                    addon_label,
                    f"{int(addon_total):,}",
                    delta=f"{take_up_rate:.1f}% take-up",
                    delta_color="off",
                )

            addon_figure = px.bar(
                addon_totals.reset_index(),
                x="Total",
                y="Add-On",
                orientation="h",
                text="Total",
                title="Total Add-On Purchases",
            )

            maximum_addon = addon_totals.max()

            addon_figure.update_traces(
                marker_color=[
                    ACCENT_COLOR
                    if value == maximum_addon
                    else NEUTRAL_COLOR
                    for value in addon_totals
                ]
            )

            addon_figure.update_layout(
                xaxis_title="Purchases",
                yaxis_title="",
            )

            st.plotly_chart(
                addon_figure,
                use_container_width=True,
            )

            st.markdown("##### Daily Add-On Table")

            st.dataframe(
                addon_summary_table,
                use_container_width=True,
            )

            st.caption(
                "Each column is a registration date; scroll "
                "horizontally for the full campaign. The Total "
                "column on the right is the grand total per "
                "add-on, and Total Add-Ons Purchased is the "
                "bottom row. Take-up rate is purchases divided "
                "by total registrations in the current "
                "selection, so a rising rate over time signals "
                "the add-on is being pitched more effectively "
                "at checkout, not just riding overall volume."
            )

            st.markdown("##### Add-Ons by Race Category")

            st.caption(
                "One pair of tables per add-on — pulled every "
                "Tuesday, showing the cumulative total as of end "
                "of the previous day (Monday), inclusive. The "
                "breakdown table splits each race category into "
                "Local and International rows with a TOTAL row "
                "underneath; the Summary table below it rolls "
                "Local and International back together per "
                "category. Country values that could not be "
                "identified are counted under International so "
                "every row still foots to the same total as the "
                "Add-On Purchases table above."
            )

            if not addon_market_category_trackers:
                st.info(
                    "No add-on purchases in a mapped category "
                    "for the current selection."
                )
            else:
                addon_tabs = st.tabs(
                    list(addon_market_category_trackers.keys())
                )

                for addon_tab, (
                    addon_label,
                    (breakdown_table, summary_table),
                ) in zip(
                    addon_tabs,
                    addon_market_category_trackers.items(),
                ):
                    with addon_tab:
                        if breakdown_table.empty:
                            st.info(
                                f"No {addon_label} purchases in "
                                "a mapped category for the "
                                "current selection."
                            )
                        else:
                            st.dataframe(
                                breakdown_table,
                                use_container_width=True,
                            )

                            st.markdown("**Summary**")

                            st.dataframe(
                                summary_table,
                                use_container_width=True,
                            )


# =========================================================
# TAB 7: REGISTRATION TIMING
# =========================================================

with timing_tab:
    st.subheader("Registration Timing Analysis")

    valid_timing_data = filtered_df.dropna(
        subset=["Registration Date Time"]
    ).copy()

    if valid_timing_data.empty:
        st.info(
            "No valid registration dates are available for timing analysis."
        )
    else:
        weekday_summary = (
            valid_timing_data[
                "Registration Weekday"
            ]
            .value_counts()
            .reindex(
                WEEKDAY_ORDER,
                fill_value=0,
            )
            .rename_axis("Weekday")
            .reset_index(name="Registrations")
        )

        weekday_figure = px.bar(
            weekday_summary,
            x="Weekday",
            y="Registrations",
            text="Registrations",
            title="Registrations by Day of Week versus Median",
        )

        # Same convention as the daily chart: days above the median
        # weekday carry the accent colour, ordinary days stay grey,
        # and a dotted line marks the median so "above" is visible
        # rather than implied.
        weekday_median = float(
            weekday_summary["Registrations"].median()
        )

        weekday_figure.update_traces(
            marker_color=[
                ACCENT_COLOR
                if value > weekday_median
                else NEUTRAL_COLOR
                for value in weekday_summary["Registrations"]
            ]
        )

        weekday_figure.add_hline(
            y=weekday_median,
            line_dash="dot",
            line_color="#3B3F46",
            annotation_text=(
                f"Median weekday {weekday_median:,.0f}"
            ),
            annotation_position="top left",
        )

        weekday_figure.update_layout(
            xaxis_title="Day of week",
            yaxis_title="Registrations",
        )

        st.plotly_chart(
            weekday_figure,
            use_container_width=True,
        )

        st.caption(
            "Blue days collect more registrations than the median "
            "weekday — these are the natural windows for campaign "
            "sends and social posts. Grey days are quieter and "
            "better suited to preparing content than launching it."
        )

        has_hour_information = (
            valid_timing_data[
                "Registration Date Time"
            ].dt.hour.ne(0).any()
            or valid_timing_data[
                "Registration Date Time"
            ].dt.minute.ne(0).any()
        )

        if has_hour_information:
            hourly_summary = (
                valid_timing_data[
                    "Registration Hour"
                ]
                .value_counts()
                .reindex(
                    range(24),
                    fill_value=0,
                )
                .rename_axis("Hour")
                .reset_index(name="Registrations")
            )

            hourly_summary["Time"] = (
                hourly_summary["Hour"]
                .map(lambda value: f"{value:02d}:00")
            )

            hourly_figure = px.bar(
                hourly_summary,
                x="Time",
                y="Registrations",
                title="Registrations by Hour of Day versus Median",
            )

            hourly_median = float(
                hourly_summary["Registrations"].median()
            )

            hourly_figure.update_traces(
                marker_color=[
                    ACCENT_COLOR
                    if value > hourly_median
                    else NEUTRAL_COLOR
                    for value in hourly_summary[
                        "Registrations"
                    ]
                ]
            )

            hourly_figure.add_hline(
                y=hourly_median,
                line_dash="dot",
                line_color="#3B3F46",
                annotation_text=(
                    f"Median hour {hourly_median:,.0f}"
                ),
                annotation_position="top left",
            )

            hourly_figure.update_layout(
                xaxis_title="Time of day",
                yaxis_title="Registrations",
            )

            st.plotly_chart(
                hourly_figure,
                use_container_width=True,
            )

            st.caption(
                "Blue hours beat the median hour. A contiguous "
                "blue block (for example lunchtime or late "
                "evening) marks the daily window when interest "
                "peaks — the heatmap below shows whether that "
                "window shifts between weekdays and weekends."
            )

            busiest_hour_row = hourly_summary.loc[
                hourly_summary[
                    "Registrations"
                ].idxmax()
            ]

            st.metric(
                "Busiest Registration Hour",
                busiest_hour_row["Time"],
                delta=(
                    f"{int(busiest_hour_row['Registrations']):,} "
                    "registrations"
                ),
                delta_color="off",
            )

            # The weekday and hourly charts hide their interaction;
            # the heatmap surfaces patterns such as weekday lunch
            # hours versus weekend mornings, which is directly
            # actionable for scheduling marketing pushes.
            st.markdown("#### When Do People Register?")

            heatmap_data = (
                valid_timing_data.groupby(
                    [
                        "Registration Weekday",
                        "Registration Hour",
                    ],
                    observed=True,
                )
                .size()
                .unstack(fill_value=0)
                .reindex(WEEKDAY_ORDER, fill_value=0)
                .reindex(
                    columns=range(24),
                    fill_value=0,
                )
            )

            heatmap_figure = go.Figure(
                data=go.Heatmap(
                    z=heatmap_data.values,
                    x=[
                        f"{hour:02d}:00"
                        for hour in heatmap_data.columns
                    ],
                    y=heatmap_data.index.tolist(),
                    colorscale=[
                        [0.0, "#F4F7FA"],
                        [0.5, "#5AA9C9"],
                        [1.0, "#16406E"],
                    ],
                    hovertemplate=(
                        "%{y} %{x}<br>Registrations: "
                        "%{z:,}<extra></extra>"
                    ),
                    colorbar={"title": "Registrations"},
                )
            )

            heatmap_figure.update_layout(
                title="Registrations by Weekday and Hour",
                xaxis_title="Hour of day",
                yaxis_title="",
                yaxis={"autorange": "reversed"},
                height=380,
            )

            st.plotly_chart(
                heatmap_figure,
                use_container_width=True,
            )

        else:
            st.info(
                "The selected registration-date field does not contain "
                "meaningful time-of-day information. Hourly analysis is hidden."
            )

        weekend_count = int(
            valid_timing_data[
                "Registration Weekday"
            ].isin(
                [
                    "Saturday",
                    "Sunday",
                ]
            ).sum()
        )

        weekday_count = (
            len(valid_timing_data) - weekend_count
        )

        timing_metric_1, timing_metric_2 = st.columns(2)

        timing_metric_1.metric(
            "Weekday Registrations",
            f"{weekday_count:,}",
        )

        timing_metric_2.metric(
            "Weekend Registrations",
            f"{weekend_count:,}",
        )


# =========================================================
# TAB 7: DATA QUALITY
# =========================================================

with quality_tab:
    st.subheader("Data Quality")

    st.dataframe(
        quality_summary,
        use_container_width=True,
        hide_index=True,
    )

    quality_figure = px.bar(
        quality_summary[
            quality_summary[
                "Data Quality Check"
            ].ne("Total source rows")
        ].sort_values("Affected Records"),
        x="Affected Records",
        y="Data Quality Check",
        orientation="h",
        text="Affected Records",
        title="Records Requiring Review",
        color_discrete_sequence=[NEUTRAL_COLOR],
    )

    quality_figure.update_layout(
        xaxis_title="Affected records",
        yaxis_title="",
    )

    st.plotly_chart(
        quality_figure,
        use_container_width=True,
    )

    unmapped_categories = (
        prepared_df[
            prepared_df[
                "Grouped Category"
            ].eq("Unmapped")
        ][
            [
                "Original Category",
                "Grouped Category",
            ]
        ]
        .drop_duplicates()
        .sort_values("Original Category")
    )

    if not unmapped_categories.empty:
        st.warning(
            "Some source categories could not be assigned to the "
            "eight summarised categories."
        )

        st.dataframe(
            unmapped_categories,
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.success(
            "All source category values were successfully assigned "
            "to the eight summarised categories."
        )

    possible_duplicate_data = prepared_df[
        prepared_df["Possible Duplicate"]
    ]

    if not possible_duplicate_data.empty:
        with st.expander(
            "Review possible duplicate rows",
            expanded=False,
        ):
            st.caption(
                "These are identical analytical rows. Without a unique "
                "participant or order ID, the application cannot confirm "
                "whether they are true duplicates."
            )

            st.dataframe(
                possible_duplicate_data,
                use_container_width=True,
                hide_index=True,
            )


# =========================================================
# TAB 8: REGISTRATION TRACKER
# =========================================================

with tracker_tab:
    st.subheader("Registration Tracker")

    tracker_mode = st.radio(
        "Tracker display",
        options=[
            "Cumulative",
            "Daily",
        ],
        horizontal=True,
    )

    if tracker_mode == "Cumulative":
        display_tracker = tracker_table

    else:
        display_tracker = create_daily_category_pivot(
            filtered_df
        )

        if not display_tracker.empty:
            display_tracker.loc[
                "Daily Registrations"
            ] = display_tracker.sum(axis=0)

    if display_tracker.empty:
        st.info(
            "No tracker information is available for the selected filters."
        )
    else:
        formatted_tracker = format_tracker_for_display(
            display_tracker
        )

        st.dataframe(
            formatted_tracker,
            use_container_width=True,
        )

    st.markdown("#### By Channel: Retail, Corporate & Complimentary")

    st.caption(
        "The same by-category tracker, split into three views by "
        "registration channel — so it's clear how much of any "
        "day's growth was organic retail demand versus a "
        "corporate batch or a complimentary allocation."
    )

    channel_view_mode = st.radio(
        "View",
        options=["Cumulative", "Daily"],
        horizontal=True,
        key="channel_tracker_view_mode",
        help=(
            "Cumulative shows the running total up to each date. "
            "Daily shows that day's new registrations only, per "
            "category, without accumulating."
        ),
    )

    active_channel_trackers = (
        channel_trackers
        if channel_view_mode == "Cumulative"
        else channel_trackers_daily
    )

    channel_tabs = st.tabs(
        list(active_channel_trackers.keys())
    )

    for channel_tab, (channel_label, channel_tracker) in zip(
        channel_tabs, active_channel_trackers.items()
    ):
        with channel_tab:
            if channel_tracker.empty:
                st.info(
                    f"No {channel_label} registrations in the "
                    "current selection."
                )
            else:
                st.dataframe(
                    format_tracker_for_display(
                        channel_tracker
                    ),
                    use_container_width=True,
                )

    st.markdown("#### Download Report")

    download_col_1, download_col_2 = st.columns(2)

    with download_col_1:
        st.download_button(
            label="Download Full Report (Excel)",
            data=excel_report_bytes,
            file_name=(
                f"{original_filename}_registration_analysis.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
            key="tracker_download_excel",
        )

    with download_col_2:
        st.download_button(
            label="Download All Tables (CSV)",
            data=archive_csv_bytes,
            file_name=(
                f"{original_filename}_archive_tables.csv"
            ),
            mime="text/csv",
            use_container_width=True,
            key="tracker_download_csv",
        )